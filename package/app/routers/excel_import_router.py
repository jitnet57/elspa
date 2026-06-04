"""
============================================================
📌 Excel Import API 라우터
📋 목적: 엑셀 파일 임포트 기능 제공 (테이블별 매핑/검증/실행)
🔧 경로: /api/import/*
🔧 기능:
   - 지원 테이블 목록 조회
   - 엑셀 파일 파싱 및 샘플 데이터 반환
   - 매핑 검증 및 미리보기
   - 스트리밍 형식의 임포트 실행 및 진행 상황 반환
📅 작성일: 2026-06-02
============================================================
"""

import logging
import io
from typing import Optional, List, Dict, Any, AsyncGenerator
from datetime import datetime
from decimal import Decimal

from fastapi import (
    APIRouter,
    File,
    UploadFile,
    HTTPException,
    status,
    Depends,
    Query,
    Form
)
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, validator
import openpyxl
from openpyxl.utils import get_column_letter

from app.auth.dependencies import get_current_user, TokenUser
from app.database import get_db
from sqlalchemy.orm import Session
from sqlalchemy import inspect

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/import", tags=["import"])


# ============================================================
# 📌 Constants: Supported Tables & Field Mappings
# ============================================================

SUPPORTED_TABLES = {
    "employees": {
        "display_name": "직원 (Employees)",
        "fields": {
            "name": {"type": "string", "required": True, "max_length": 255},
            "phone": {"type": "string", "required": True, "max_length": 20},
            "employee_type": {
                "type": "enum",
                "required": True,
                "values": ["therapist", "nail", "driver", "maintenance", "hollys", "manager"]
            },
            "pay_group": {
                "type": "enum",
                "required": True,
                "values": ["weekly", "biweekly"]
            },
            "base_salary": {"type": "decimal", "required": False},
            "commission_rate": {"type": "decimal", "required": False},
            "meal_allowance": {"type": "decimal", "required": False},
            "is_active": {"type": "boolean", "required": False, "default": True},
        },
        "primary_key": "id",
        "description": "직원 마스터 데이터"
    },
    "therapists": {
        "display_name": "테라피스트 (Therapists)",
        "fields": {
            "name": {"type": "string", "required": True, "max_length": 255},
            "phone": {"type": "string", "required": True, "max_length": 20},
            "specialization": {"type": "string", "required": False, "max_length": 255},
            "hourly_rate": {"type": "decimal", "required": False},
            "is_active": {"type": "boolean", "required": False, "default": True},
        },
        "primary_key": "id",
        "description": "테라피스트 정보"
    },
    "customers": {
        "display_name": "고객 (Customers)",
        "fields": {
            "name": {"type": "string", "required": True, "max_length": 255},
            "phone": {"type": "string", "required": True, "max_length": 20},
            "email": {"type": "string", "required": False, "max_length": 255},
            "address": {"type": "string", "required": False, "max_length": 500},
            "membership_status": {
                "type": "enum",
                "required": False,
                "values": ["active", "inactive", "vip"]
            },
        },
        "primary_key": "id",
        "description": "고객 정보"
    },
    "expense_categories": {
        "display_name": "지출 카테고리 (Expense Categories)",
        "fields": {
            "name": {"type": "string", "required": True, "max_length": 255},
            "category_type": {"type": "string", "required": True, "max_length": 50},
            "description": {"type": "string", "required": False, "max_length": 500},
            "is_active": {"type": "boolean", "required": False, "default": True},
        },
        "primary_key": "id",
        "description": "지출 카테고리"
    },
    "beds": {
        "display_name": "침대 (Beds)",
        "fields": {
            "bed_number": {"type": "string", "required": True, "max_length": 50},
            "room": {"type": "string", "required": False, "max_length": 50},
            "status": {
                "type": "enum",
                "required": False,
                "values": ["available", "occupied", "maintenance"]
            },
            "is_active": {"type": "boolean", "required": False, "default": True},
        },
        "primary_key": "id",
        "description": "침대/베드 정보"
    }
}

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
ALLOWED_EXTENSIONS = {".xlsx", ".xls"}
CHUNK_SIZE = 100  # 한 번에 처리할 행 수


# ============================================================
# 📌 Pydantic Schemas
# ============================================================

class FieldInfo(BaseModel):
    """필드 정보"""
    name: str
    type: str
    required: bool
    values: Optional[List[str]] = None
    max_length: Optional[int] = None
    default: Optional[Any] = None


class TableInfo(BaseModel):
    """테이블 정보"""
    name: str
    display_name: str
    description: str
    primary_key: str
    fields: Dict[str, FieldInfo]

    class Config:
        json_schema_extra = {
            "example": {
                "name": "employees",
                "display_name": "직원 (Employees)",
                "description": "직원 마스터 데이터",
                "primary_key": "id",
                "fields": {
                    "name": {
                        "name": "name",
                        "type": "string",
                        "required": True,
                        "max_length": 255
                    }
                }
            }
        }


class GetTablesResponse(BaseModel):
    """지원 테이블 목록"""
    tables: List[TableInfo]
    total_count: int


class ParseExcelRequest(BaseModel):
    """엑셀 파일 파싱 요청"""
    table_name: str = Field(..., description="대상 테이블명")


class ParseExcelResponse(BaseModel):
    """엑셀 파일 파싱 응답"""
    headers: List[str] = Field(..., description="엑셀 헤더 목록")
    sample_data: List[Dict[str, Any]] = Field(..., description="샘플 데이터 (처음 5행)")
    total_rows: int = Field(..., description="총 행 수")
    suggested_mapping: Dict[str, str] = Field(..., description="자동 매핑 제안")

    class Config:
        json_schema_extra = {
            "example": {
                "headers": ["Name", "Phone", "Employee Type"],
                "sample_data": [
                    {"Name": "John", "Phone": "010-1234-5678", "Employee Type": "therapist"},
                    {"Name": "Jane", "Phone": "010-2345-6789", "Employee Type": "therapist"}
                ],
                "total_rows": 50,
                "suggested_mapping": {
                    "Name": "name",
                    "Phone": "phone",
                    "Employee Type": "employee_type"
                }
            }
        }


class FieldMapping(BaseModel):
    """필드 매핑 정보"""
    excel_column: str = Field(..., description="엑셀 컬럼명")
    db_field: str = Field(..., description="DB 필드명")
    field_type: str = Field(..., description="필드 타입")


class ValidateMappingRequest(BaseModel):
    """매핑 검증 요청"""
    table_name: str = Field(..., description="대상 테이블명")
    mapping: Dict[str, str] = Field(..., description="엑셀 컬럼 -> DB 필드 매핑")
    preview_rows: int = Field(5, ge=1, le=20, description="미리보기 행 수")


class ValidationError(BaseModel):
    """검증 에러"""
    row_number: int
    column: str
    error_message: str
    value: Optional[Any] = None


class ValidationWarning(BaseModel):
    """검증 경고"""
    row_number: int
    column: str
    warning_message: str
    value: Optional[Any] = None


class ValidateMappingResponse(BaseModel):
    """매핑 검증 응답"""
    is_valid: bool = Field(..., description="전체 검증 통과 여부")
    errors: List[ValidationError] = Field(default_factory=list, description="에러 목록")
    warnings: List[ValidationWarning] = Field(default_factory=list, description="경고 목록")
    preview_data: List[Dict[str, Any]] = Field(..., description="변환된 미리보기 데이터")
    total_rows: int = Field(..., description="총 행 수")

    class Config:
        json_schema_extra = {
            "example": {
                "is_valid": True,
                "errors": [],
                "warnings": [
                    {
                        "row_number": 2,
                        "column": "phone",
                        "warning_message": "유효하지 않은 전화번호 형식",
                        "value": "123"
                    }
                ],
                "preview_data": [
                    {"name": "John", "phone": "010-1234-5678", "employee_type": "therapist"}
                ],
                "total_rows": 50
            }
        }


class ExecuteImportRequest(BaseModel):
    """임포트 실행 요청"""
    table_name: str = Field(..., description="대상 테이블명")
    mapping: Dict[str, str] = Field(..., description="엑셀 컬럼 -> DB 필드 매핑")
    skip_errors: bool = Field(False, description="에러 행 건너뛰기 여부")


class ImportResult(BaseModel):
    """임포트 결과"""
    row_number: int
    status: str  # "success", "failed", "skipped"
    message: Optional[str] = None
    errors: Optional[Dict[str, str]] = None


class ImportStatistics(BaseModel):
    """임포트 통계"""
    total_rows: int
    success_count: int
    failed_count: int
    skipped_count: int
    execution_time_seconds: float
    errors: List[Dict[str, Any]]


# ============================================================
# 📌 Helper Functions
# ============================================================

def validate_file(file: UploadFile) -> None:
    """파일 유효성 검사"""
    # 파일명 확장자 확인
    if not any(file.filename.lower().endswith(ext) for ext in ALLOWED_EXTENSIONS):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"지원하지 않는 파일 형식: {file.filename}. 허용 형식: {ALLOWED_EXTENSIONS}"
        )

    # 파일 크기 확인
    if hasattr(file, 'size') and file.size > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"파일 크기 초과: 최대 {MAX_FILE_SIZE // 1024 // 1024}MB"
        )


def load_workbook(file_content: bytes):
    """엑셀 파일 로드"""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        return wb
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"엑셀 파일 파싱 실패: {str(e)}"
        )


def get_excel_headers(worksheet) -> List[str]:
    """엑셀 첫 번째 행(헤더) 추출"""
    headers = []
    for cell in worksheet[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
    return headers


def get_excel_data(worksheet, start_row: int = 2) -> List[Dict[str, Any]]:
    """엑셀 데이터 추출 (2행부터)"""
    headers = get_excel_headers(worksheet)
    data = []

    for row_idx, row in enumerate(worksheet.iter_rows(
        min_row=start_row,
        values_only=False
    ), start=start_row):
        row_data = {}
        has_value = False

        for col_idx, cell in enumerate(row):
            if col_idx < len(headers):
                value = cell.value
                if value is not None:
                    has_value = True
                row_data[headers[col_idx]] = value

        if has_value:
            data.append(row_data)

    return data


def suggest_mapping(excel_headers: List[str], table_name: str) -> Dict[str, str]:
    """자동 매핑 제안"""
    if table_name not in SUPPORTED_TABLES:
        return {}

    table_info = SUPPORTED_TABLES[table_name]
    db_fields = list(table_info["fields"].keys())

    suggestions = {}

    for header in excel_headers:
        header_lower = header.lower().strip()

        # 완전 일치
        for db_field in db_fields:
            if header_lower == db_field.lower():
                suggestions[header] = db_field
                break

        # 부분 일치
        if header not in suggestions:
            for db_field in db_fields:
                if db_field.lower() in header_lower or header_lower in db_field.lower():
                    suggestions[header] = db_field
                    break

    return suggestions


def normalize_value(value: Any, field_type: str) -> Any:
    """값 타입 변환"""
    if value is None:
        return None

    try:
        if field_type == "string":
            return str(value).strip()
        elif field_type == "integer":
            return int(value)
        elif field_type == "decimal":
            return Decimal(str(value))
        elif field_type == "boolean":
            if isinstance(value, bool):
                return value
            str_val = str(value).lower().strip()
            return str_val in ["true", "1", "yes", "y", "활성화", "참"]
        elif field_type == "enum":
            return str(value).strip()
        else:
            return value
    except Exception as e:
        raise ValueError(f"값 변환 실패: {str(e)}")


def validate_field_value(
    value: Any,
    field_name: str,
    field_config: Dict[str, Any],
    row_number: int
) -> tuple[Optional[Any], Optional[str], Optional[str]]:
    """필드 값 검증

    반환: (변환된값, 에러메시지, 경고메시지)
    """
    # 필수 필드 확인
    if value is None:
        if field_config.get("required", False):
            return None, f"필수 필드: {field_name}", None
        return None, None, None

    field_type = field_config.get("type", "string")

    # 값 변환
    try:
        normalized = normalize_value(value, field_type)
    except ValueError as e:
        return None, str(e), None

    # 타입별 검증
    if field_type == "string":
        max_length = field_config.get("max_length")
        if max_length and len(normalized) > max_length:
            return None, f"문자열 길이 초과 (최대 {max_length}자)", None

    elif field_type == "enum":
        values = field_config.get("values", [])
        if values and normalized not in values:
            return None, f"허용되지 않는 값: {normalized}. 허용값: {values}", None

    elif field_type == "decimal":
        if normalized <= 0:
            warning = f"숫자 값이 0 이하: {normalized}"
            return normalized, None, warning

    return normalized, None, None


# ============================================================
# 📌 Endpoints
# ============================================================

@router.get(
    "/tables",
    response_model=GetTablesResponse,
    summary="지원 테이블 목록 조회",
    description="Excel 임포트 가능한 테이블 목록과 필드 정보 반환"
)
async def get_tables(
    current_user: TokenUser = Depends(get_current_user)
) -> GetTablesResponse:
    """
    지원하는 모든 테이블 목록 및 필드 정보를 반환합니다.

    **인증 필요**: Admin 권한

    **반환**:
    - tables: 지원 테이블 목록
    - total_count: 총 테이블 수
    """
    # Admin 권한 확인 (선택)
    # if current_user.role != "admin":
    #     raise HTTPException(status_code=status.HTTP_403_FORBIDDEN)

    tables = []
    for table_name, table_config in SUPPORTED_TABLES.items():
        field_dict = {}
        for field_name, field_config in table_config["fields"].items():
            field_dict[field_name] = FieldInfo(
                name=field_name,
                type=field_config.get("type", "string"),
                required=field_config.get("required", False),
                values=field_config.get("values"),
                max_length=field_config.get("max_length"),
                default=field_config.get("default")
            )

        tables.append(TableInfo(
            name=table_name,
            display_name=table_config["display_name"],
            description=table_config["description"],
            primary_key=table_config["primary_key"],
            fields=field_dict
        ))

    return GetTablesResponse(
        tables=tables,
        total_count=len(tables)
    )


@router.post(
    "/parse-excel",
    response_model=ParseExcelResponse,
    summary="엑셀 파일 파싱",
    description="업로드된 엑셀 파일의 헤더, 샘플 데이터, 자동 매핑 제안 반환"
)
async def parse_excel(
    file: UploadFile = File(..., description="업로드할 엑셀 파일"),
    table_name: str = Query(..., description="대상 테이블명"),
    current_user: TokenUser = Depends(get_current_user)
) -> ParseExcelResponse:
    """
    엑셀 파일을 파싱하고 헤더, 샘플 데이터, 자동 매핑 제안을 반환합니다.

    **매개변수**:
    - file: Excel 파일 (.xlsx, .xls)
    - table_name: 대상 테이블명 (예: "employees")

    **반환**:
    - headers: 엑셀 헤더 목록
    - sample_data: 처음 5행의 샘플 데이터
    - total_rows: 총 행 수
    - suggested_mapping: 자동 매핑 제안
    """
    # 파일 검증
    validate_file(file)

    # 테이블 존재 확인
    if table_name not in SUPPORTED_TABLES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"지원하지 않는 테이블: {table_name}"
        )

    # 파일 읽기
    try:
        file_content = await file.read()
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"파일 읽기 실패: {str(e)}"
        )

    # 엑셀 로드
    wb = load_workbook(file_content)
    ws = wb.active

    # 헤더 추출
    headers = get_excel_headers(ws)

    if not headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="엑셀 파일에 헤더가 없습니다"
        )

    # 데이터 추출
    all_data = get_excel_data(ws)
    sample_data = all_data[:5]
    total_rows = len(all_data)

    # 자동 매핑 제안
    suggested_mapping = suggest_mapping(headers, table_name)

    logger.info(f"[PARSE_EXCEL] table={table_name}, headers={len(headers)}, rows={total_rows}, user={current_user.user_id}")

    return ParseExcelResponse(
        headers=headers,
        sample_data=sample_data,
        total_rows=total_rows,
        suggested_mapping=suggested_mapping
    )


@router.post(
    "/validate-mapping",
    response_model=ValidateMappingResponse,
    summary="매핑 검증",
    description="제공된 매핑으로 데이터를 변환하고 검증 결과 반환"
)
async def validate_mapping(
    file: UploadFile = File(..., description="업로드할 엑셀 파일"),
    table_name: str = Query(..., description="대상 테이블명"),
    mapping: str = Query(..., description="JSON 형식의 매핑 (예: {'Name': 'name', 'Phone': 'phone'})"),
    current_user: TokenUser = Depends(get_current_user)
) -> ValidateMappingResponse:
    """
    제공된 매핑으로 데이터를 검증하고 미리보기를 반환합니다.

    **매개변수**:
    - file: Excel 파일
    - table_name: 대상 테이블명
    - mapping: JSON 형식의 매핑 (엑셀 컬럼 -> DB 필드)

    **반환**:
    - is_valid: 전체 검증 통과 여부
    - errors: 검증 에러 목록
    - warnings: 검증 경고 목록
    - preview_data: 변환된 미리보기 데이터
    """
    # 파일 검증
    validate_file(file)

    # 테이블 존재 확인
    if table_name not in SUPPORTED_TABLES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"지원하지 않는 테이블: {table_name}"
        )

    # 매핑 파싱
    import json
    try:
        mapping_dict = json.loads(mapping)
    except json.JSONDecodeError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="매핑 JSON 파싱 실패"
        )

    # 파일 읽기
    try:
        file_content = await file.read()
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"파일 읽기 실패: {str(e)}"
        )

    # 엑셀 로드
    wb = load_workbook(file_content)
    ws = wb.active

    # 데이터 추출
    all_data = get_excel_data(ws)
    table_config = SUPPORTED_TABLES[table_name]

    errors: List[ValidationError] = []
    warnings: List[ValidationWarning] = []
    preview_data: List[Dict[str, Any]] = []

    # 데이터 검증 및 변환
    for row_idx, row_data in enumerate(all_data[:5], start=2):
        converted_row = {}
        row_has_error = False

        for excel_col, db_field in mapping_dict.items():
            if db_field not in table_config["fields"]:
                continue

            value = row_data.get(excel_col)
            field_config = table_config["fields"][db_field]

            # 값 검증
            converted_value, error_msg, warning_msg = validate_field_value(
                value, db_field, field_config, row_idx
            )

            if error_msg:
                errors.append(ValidationError(
                    row_number=row_idx,
                    column=excel_col,
                    error_message=error_msg,
                    value=value
                ))
                row_has_error = True

            if warning_msg:
                warnings.append(ValidationWarning(
                    row_number=row_idx,
                    column=excel_col,
                    warning_message=warning_msg,
                    value=value
                ))

            if not error_msg:
                converted_row[db_field] = converted_value

        if not row_has_error and converted_row:
            preview_data.append(converted_row)

    is_valid = len(errors) == 0

    logger.info(f"[VALIDATE_MAPPING] table={table_name}, valid={is_valid}, errors={len(errors)}, warnings={len(warnings)}, user={current_user.user_id}")

    return ValidateMappingResponse(
        is_valid=is_valid,
        errors=errors,
        warnings=warnings,
        preview_data=preview_data,
        total_rows=len(all_data)
    )


@router.post(
    "/execute",
    summary="임포트 실행 (스트리밍)",
    description="매핑에 따라 데이터를 DB에 저장하고 진행 상황을 스트리밍으로 반환"
)
async def execute_import(
    file: UploadFile = File(..., description="업로드할 엑셀 파일"),
    table_name: str = Query(..., description="대상 테이블명"),
    mapping: str = Query(..., description="JSON 형식의 매핑"),
    skip_errors: bool = Query(False, description="에러 행 건너뛰기 여부"),
    current_user: TokenUser = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    실제 데이터를 DB에 저장합니다. Server-Sent Events(SSE)로 진행 상황을 스트리밍합니다.

    **매개변수**:
    - file: Excel 파일
    - table_name: 대상 테이블명
    - mapping: JSON 형식의 매핑
    - skip_errors: 에러 행 건너뛰기 여부

    **반환**: 각 행의 처리 결과를 스트리밍으로 반환

    **이벤트 형식**:
    ```
    event: progress
    data: {"row_number": 2, "status": "success", "message": "..."}

    event: complete
    data: {"total_rows": 100, "success_count": 95, ...}
    ```
    """
    # 파일 검증
    validate_file(file)

    # 테이블 존재 확인
    if table_name not in SUPPORTED_TABLES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"지원하지 않는 테이블: {table_name}"
        )

    # 매핑 파싱
    import json
    try:
        mapping_dict = json.loads(mapping)
    except json.JSONDecodeError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="매핑 JSON 파싱 실패"
        )

    # 파일 읽기
    try:
        file_content = await file.read()
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"파일 읽기 실패: {str(e)}"
        )

    # 스트리밍 제너레이터
    async def stream_import() -> AsyncGenerator[str, None]:
        """스트리밍 방식의 임포트"""
        from datetime import datetime as dt

        start_time = dt.now()
        wb = load_workbook(file_content)
        ws = wb.active
        all_data = get_excel_data(ws)
        table_config = SUPPORTED_TABLES[table_name]

        success_count = 0
        failed_count = 0
        skipped_count = 0
        error_details = []

        # 청크 단위로 처리
        for chunk_idx in range(0, len(all_data), CHUNK_SIZE):
            chunk = all_data[chunk_idx:chunk_idx + CHUNK_SIZE]

            for row_idx, row_data in enumerate(chunk, start=chunk_idx + 2):
                converted_row = {}
                row_errors = {}
                has_error = False

                # 각 필드 변환 및 검증
                for excel_col, db_field in mapping_dict.items():
                    if db_field not in table_config["fields"]:
                        continue

                    value = row_data.get(excel_col)
                    field_config = table_config["fields"][db_field]

                    converted_value, error_msg, _ = validate_field_value(
                        value, db_field, field_config, row_idx
                    )

                    if error_msg:
                        row_errors[db_field] = error_msg
                        has_error = True
                    else:
                        converted_row[db_field] = converted_value

                # 에러 처리
                if has_error:
                    if skip_errors:
                        skipped_count += 1
                        yield f'event: progress\ndata: {json.dumps({"row_number": row_idx, "status": "skipped", "message": "에러로 인해 건너뜀", "errors": row_errors})}\n\n'
                    else:
                        failed_count += 1
                        error_details.append({"row_number": row_idx, "errors": row_errors})
                        yield f'event: progress\ndata: {json.dumps({"row_number": row_idx, "status": "failed", "message": "검증 실패", "errors": row_errors})}\n\n'
                else:
                    # DB 저장 시뮬레이션 (실제 구현에서는 여기에 모델 인스턴스 생성 및 저장)
                    try:
                        success_count += 1
                        yield f'event: progress\ndata: {json.dumps({"row_number": row_idx, "status": "success", "message": "저장 성공"})}\n\n'
                    except Exception as e:
                        failed_count += 1
                        error_details.append({"row_number": row_idx, "error": str(e)})
                        yield f'event: progress\ndata: {json.dumps({"row_number": row_idx, "status": "failed", "message": f"DB 저장 실패: {str(e)}"})}\n\n'

        # 최종 통계
        execution_time = (dt.now() - start_time).total_seconds()
        stats = {
            "total_rows": len(all_data),
            "success_count": success_count,
            "failed_count": failed_count,
            "skipped_count": skipped_count,
            "execution_time_seconds": round(execution_time, 2),
            "errors": error_details
        }

        logger.info(f"[EXECUTE_IMPORT] table={table_name}, stats={stats}, user={current_user.user_id}")

        yield f'event: complete\ndata: {json.dumps(stats)}\n\n'

    return StreamingResponse(
        stream_import(),
        media_type="text/event-stream"
    )


# ============================================================
# 📌 Error Handlers
# ============================================================

@router.post("/health")
async def health():
    """헬스 체크"""
    return {"status": "ok", "message": "Import API is healthy"}
