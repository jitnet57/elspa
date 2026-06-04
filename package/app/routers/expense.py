"""
Daily Expense Report — Receipt Scan + DB CRUD + Excel
GET  /api/expense/dates          : 저장된 날짜 목록
GET  /api/expense/records        : 날짜별 지출 목록 (?date=2026-05-21)
POST /api/expense/scan           : 영수증 OCR → DB 저장 → JSON 반환
POST /api/expense/records        : 지출 1건 추가
PUT  /api/expense/records/{id}   : 지출 수정
DELETE /api/expense/records/{id} : 지출 삭제
POST /api/expense/export         : 날짜별 DB 데이터 → Excel 다운로드
"""

import io, base64, json, logging
from typing import List, Optional
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, UploadFile, File, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func as sqlfunc
import anthropic

from app.config import settings
from app.database import get_db
from app.models.financial import Expense

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/expense", tags=["expense"])

ALLOWED_TYPES = {"image/jpeg", "image/jpg", "image/png", "image/webp", "image/gif"}

CATEGORIES = {
    "food":          "🍽️ Food (식비)",
    "transport":     "🚗 Transport (교통비)",
    "supplies":      "🛒 Supplies (소모품)",
    "utilities":     "💡 Utilities (공과금)",
    "entertainment": "🤝 Entertainment (접대비)",
    "medical":       "🏥 Medical (의료비)",
    "other":         "📦 Other (기타)",
}
CATEGORY_COLORS = {
    "food": "FFD9B3", "transport": "B3E5FC", "supplies": "C8E6C9",
    "utilities": "FFF9C4", "entertainment": "F8BBD9",
    "medical": "E1BEE7", "other": "EEEEEE",
}

# ── Pydantic 스키마 ───────────────────────────────────────────────

class ExpenseCreate(BaseModel):
    report_date:   str
    vendor:        Optional[str] = None
    expense_date:  Optional[str] = None
    amount:        float = 0
    currency:      str = "PHP"
    category_name: str = "other"
    items_json:    Optional[str] = None   # JSON 배열 문자열
    description:   Optional[str] = None

class ExpenseUpdate(BaseModel):
    vendor:        Optional[str]   = None
    expense_date:  Optional[str]   = None
    amount:        Optional[float] = None
    currency:      Optional[str]   = None
    category_name: Optional[str]   = None
    items_json:    Optional[str]   = None
    description:   Optional[str]   = None

class ExportRequest(BaseModel):
    report_date: str

def _to_dict(row: Expense) -> dict:
    items = []
    if row.items_json:
        try: items = json.loads(row.items_json)
        except Exception: pass
    return {
        "id":            row.id,
        "report_date":   row.report_date,
        "vendor":        row.vendor,
        "expense_date":  row.expense_date.strftime("%Y-%m-%d") if row.expense_date else None,
        "amount":        float(row.amount or 0),
        "currency":      row.currency or "PHP",
        "category_name": row.category_name or "other",
        "items":         items,
        "description":   row.description,
        "created_at":    row.created_at.isoformat() if row.created_at else None,
        "updated_at":    row.updated_at.isoformat() if row.updated_at else None,
    }

# ── OCR 프롬프트 ──────────────────────────────────────────────────

EXTRACT_PROMPT = """
You are an OCR assistant reading a receipt or expense document (any language/country).

Extract data and return ONLY valid JSON:

{
  "vendor": "store name",
  "date": "YYYY-MM-DD if possible",
  "total_amount": 123.45,
  "currency": "PHP",
  "items": ["item 1", "item 2"],
  "category_guess": "food",
  "notes": "receipt no, address, etc"
}

category_guess must be one of: food, transport, supplies, utilities, entertainment, medical, other
- food: restaurant, cafe, grocery, delivery
- transport: gas, taxi, Grab, parking, toll, bus, ferry
- supplies: office supplies, cleaning materials, hardware
- utilities: electricity, water, internet, phone, rent
- entertainment: team meals, client gifts, events
- medical: medicine, clinic, hospital
- other: anything else

total_amount must be a number. Use 0 if unreadable.
Return ONLY the JSON, no markdown.
"""

async def _ocr(image_bytes: bytes, media_type: str, filename: str) -> dict:
    client = anthropic.Anthropic(api_key=settings.anthropic_api_key)
    b64    = base64.standard_b64encode(image_bytes).decode()
    resp   = client.messages.create(
        model="claude-opus-4-5", max_tokens=1024,
        messages=[{"role": "user", "content": [
            {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
            {"type": "text",  "text": EXTRACT_PROMPT},
        ]}],
    )
    raw = resp.content[0].text.strip()
    if "```" in raw:
        raw = raw.split("```")[1]
        if raw.startswith("json"): raw = raw[4:]
    data = json.loads(raw)
    data["filename"] = filename
    return data

# ── Excel 빌더 ────────────────────────────────────────────────────

def _build_excel(report_date: str, rows: list) -> bytes:
    PESO = '#,##0.00'
    thin = Side(style="thin", color="CCCCCC")
    med  = Side(style="medium", color="555555")
    BT = Border(left=thin, right=thin, top=thin, bottom=thin)
    BM = Border(left=med,  right=med,  top=med,  bottom=med)
    BX = Border(left=thin, right=thin, top=med,  bottom=med)

    def st(ws, r, c, val=None, bold=False, size=10, color="000000",
           bg=None, ha="left", bdr=BT, fmt=None, wrap=False):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font      = Font(name="Calibri", bold=bold, size=size, color=color)
        cell.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
        if bg:  cell.fill = PatternFill("solid", fgColor=bg)
        if bdr: cell.border = bdr
        if fmt: cell.number_format = fmt
        return cell

    wb = openpyxl.Workbook()

    # ══ Sheet 1: Daily Expenses ═══════════════════════════════
    ws = wb.active
    ws.title = f"Expenses {report_date}"
    for i, w in enumerate([5, 22, 12, 26, 14, 32, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:G1")
    st(ws, 1, 1, f"DAILY EXPENSE REPORT — {report_date}",
       bold=True, size=13, color="FFFFFF", bg="1B5E20", ha="center", bdr=BM)
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:G2")
    st(ws, 2, 1, f"Generated: {datetime.now():%Y-%m-%d %H:%M}  |  ElSpa Management System",
       size=9, bg="C8E6C9", ha="center")
    ws.row_dimensions[3].height = 5

    for ci, h in enumerate(["No.", "Vendor", "Date", "Category", "Amount (PHP)", "Items", "Notes"], 1):
        st(ws, 4, ci, h, bold=True, size=10, color="FFFFFF", bg="2E7D32", ha="center", bdr=BM, wrap=True)
    ws.row_dimensions[4].height = 22

    grand = 0
    for idx, e in enumerate(rows):
        row = idx + 5
        rbg = "F1F8E9" if idx % 2 == 0 else "FFFFFF"
        amt = float(e.get("amount", 0) or 0)
        grand += amt
        cat   = e.get("category_name", "other") or "other"
        clbl  = CATEGORIES.get(cat, "📦 Other")
        cbg   = CATEGORY_COLORS.get(cat, "EEEEEE")
        items = ", ".join(e.get("items") or [])

        st(ws, row, 1, idx + 1,                   size=9, ha="center", bg=rbg)
        st(ws, row, 2, e.get("vendor") or "—",    size=9, ha="left",   bg=rbg)
        st(ws, row, 3, e.get("expense_date") or "", size=9, ha="center", bg=rbg)
        st(ws, row, 4, clbl,                      size=9, ha="left",   bg=cbg)
        st(ws, row, 5, amt,                       size=9, ha="right",  bg=rbg, fmt=PESO)
        st(ws, row, 6, items,                     size=9, ha="left",   bg=rbg, wrap=True)
        st(ws, row, 7, e.get("description") or "", size=9, ha="left",   bg=rbg)
        ws.row_dimensions[row].height = 15

    tr = 5 + len(rows)
    ws.merge_cells(f"A{tr}:D{tr}")
    st(ws, tr, 1, f"TOTAL  ({len(rows)} receipts)", bold=True, size=11,
       color="FFFFFF", bg="2E7D32", ha="center", bdr=BX)
    st(ws, tr, 5, grand, bold=True, size=11, bg="E8F5E9", ha="right", fmt=PESO, bdr=BX)
    for ci in [6, 7]: st(ws, tr, ci, "", bg="E8F5E9", bdr=BX)
    ws.row_dimensions[tr].height = 20
    ws.freeze_panes = "A5"

    # ══ Sheet 2: Category Summary ═════════════════════════════
    ws2 = wb.create_sheet("Category Summary")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 16

    ws2.merge_cells("A1:D1")
    st(ws2, 1, 1, f"CATEGORY SUMMARY — {report_date}",
       bold=True, size=13, color="FFFFFF", bg="1B5E20", ha="center", bdr=BM)

    for ci, h in enumerate(["Category", "Count", "Total (PHP)", "% of Total"], 1):
        st(ws2, 2, ci, h, bold=True, size=10, color="FFFFFF", bg="2E7D32", ha="center", bdr=BM)

    cat_data: dict = {}
    for e in rows:
        cat = e.get("category_name", "other") or "other"
        cat_data.setdefault(cat, {"count": 0, "total": 0.0})
        cat_data[cat]["count"] += 1
        cat_data[cat]["total"] += float(e.get("amount", 0) or 0)

    for ri, (cat, d) in enumerate(
        sorted(cat_data.items(), key=lambda x: x[1]["total"], reverse=True), 3
    ):
        pct = d["total"] / grand * 100 if grand else 0
        bg  = CATEGORY_COLORS.get(cat, "EEEEEE")
        st(ws2, ri, 1, CATEGORIES.get(cat, cat), size=10, bg=bg)
        st(ws2, ri, 2, d["count"],               size=10, bg=bg, ha="center")
        st(ws2, ri, 3, d["total"],               size=10, bg=bg, ha="right", fmt=PESO)
        st(ws2, ri, 4, f"{pct:.1f}%",            size=10, bg=bg, ha="center")

    tr2 = 3 + len(cat_data)
    for ci, (v, fmt) in enumerate(
        [("GRAND TOTAL", None), (len(rows), None), (grand, PESO), ("100%", None)], 1
    ):
        st(ws2, tr2, ci, v, bold=True, size=11, color="FFFFFF",
           bg="1B5E20", ha="right" if ci == 3 else "center", fmt=fmt, bdr=BM)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

# ── 엔드포인트 ────────────────────────────────────────────────────

@router.get("/dates")
async def list_dates(db: AsyncSession = Depends(get_db)):
    """저장된 날짜 목록 + 건수/합계"""
    result = await db.execute(
        select(
            Expense.report_date,
            sqlfunc.count(Expense.id).label("count"),
            sqlfunc.sum(Expense.amount).label("total"),
        )
        .where(Expense.report_date.is_not(None))
        .group_by(Expense.report_date)
        .order_by(Expense.report_date.desc())
    )
    return [
        {"date": r.report_date, "count": r.count, "total": float(r.total or 0)}
        for r in result.all()
    ]


@router.get("/range")
async def get_range(
    start: str = Query(..., description="YYYY-MM-DD"),
    end:   str = Query(..., description="YYYY-MM-DD"),
    db: AsyncSession = Depends(get_db),
):
    """기간별 비용 자동 집계 (경영지표 일간·주간 보고서용)

    report_date 는 'YYYY-MM-DD' 문자열이라 사전식 비교 = 날짜 비교가 성립한다.
    반환: 총합 / 카테고리별 합계 / 일자별 합계
    """
    cond = (Expense.report_date.is_not(None)) & \
           (Expense.report_date >= start) & (Expense.report_date <= end)

    # 일자별 합계
    date_rows = (await db.execute(
        select(Expense.report_date, sqlfunc.sum(Expense.amount).label("total"))
        .where(cond)
        .group_by(Expense.report_date)
        .order_by(Expense.report_date)
    )).all()
    by_date = [{"date": r.report_date, "total": float(r.total or 0)} for r in date_rows]

    # 카테고리별 합계
    cat_rows = (await db.execute(
        select(Expense.category_name, sqlfunc.sum(Expense.amount).label("total"))
        .where(cond)
        .group_by(Expense.category_name)
    )).all()
    by_category = {(r.category_name or "other"): float(r.total or 0) for r in cat_rows}

    return {
        "start": start,
        "end": end,
        "total": sum(by_category.values()),
        "by_category": by_category,
        "by_date": by_date,
    }


@router.get("/records")
async def get_records(
    date: str = Query(..., description="YYYY-MM-DD"),
    db: AsyncSession = Depends(get_db),
):
    """날짜별 지출 목록"""
    result = await db.execute(
        select(Expense)
        .where(Expense.report_date == date)
        .order_by(Expense.id)
    )
    return [_to_dict(r) for r in result.scalars().all()]


@router.post("/records", status_code=201)
async def create_record(data: ExpenseCreate, db: AsyncSession = Depends(get_db)):
    """지출 1건 추가"""
    exp_date = None
    if data.expense_date:
        try: exp_date = datetime.strptime(data.expense_date, "%Y-%m-%d")
        except ValueError: pass
    row = Expense(
        report_date   = data.report_date,
        vendor        = data.vendor,
        expense_date  = exp_date or datetime.utcnow(),
        amount        = data.amount,
        currency      = data.currency,
        category_name = data.category_name,
        items_json    = data.items_json,
        description   = data.description,
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return _to_dict(row)


@router.put("/records/{record_id}")
async def update_record(
    record_id: int,
    data: ExpenseUpdate,
    db: AsyncSession = Depends(get_db),
):
    """지출 수정"""
    result = await db.execute(select(Expense).where(Expense.id == record_id))
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="레코드 없음")

    for field, value in data.model_dump(exclude_none=True).items():
        if field == "expense_date" and isinstance(value, str):
            try: value = datetime.strptime(value, "%Y-%m-%d")
            except ValueError: continue
        setattr(row, field, value)

    row.updated_at = datetime.utcnow()
    await db.commit()
    await db.refresh(row)
    return _to_dict(row)


@router.delete("/records/{record_id}")
async def delete_record(record_id: int, db: AsyncSession = Depends(get_db)):
    """지출 삭제"""
    result = await db.execute(select(Expense).where(Expense.id == record_id))
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="레코드 없음")
    await db.delete(row)
    await db.commit()
    return {"success": True, "deleted_id": record_id}


@router.post("/scan")
async def scan_receipts(
    files: List[UploadFile] = File(...),
    report_date: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    """영수증 OCR → DB 저장 → JSON 반환"""
    if not files:
        raise HTTPException(status_code=400, detail="파일 없음")
    for f in files:
        if f.content_type not in ALLOWED_TYPES:
            raise HTTPException(status_code=400, detail=f"지원하지 않는 형식: {f.filename}")

    today = report_date or datetime.now().strftime("%Y-%m-%d")
    results, errors = [], []

    for i, upload in enumerate(files):
        try:
            content = await upload.read()
            data    = await _ocr(content, upload.content_type or "image/jpeg", upload.filename or "")
            amt     = float(data.get("total_amount") or 0)
            items   = data.get("items") or []
            exp_date_str = data.get("date") or today
            try:    exp_dt = datetime.strptime(exp_date_str[:10], "%Y-%m-%d")
            except: exp_dt = datetime.now()

            row = Expense(
                report_date   = today,
                vendor        = data.get("vendor"),
                expense_date  = exp_dt,
                amount        = amt,
                currency      = data.get("currency", "PHP"),
                category_name = data.get("category_guess", "other") or "other",
                items_json    = json.dumps(items),
                description   = data.get("notes"),
            )
            db.add(row)
            await db.flush()   # id 확보
            result_dict = _to_dict(row)
            result_dict["id_temp"] = i + 1
            results.append(result_dict)
            logger.info(f"✅ OCR: {upload.filename} | {data.get('vendor')} | {amt:.2f}")
        except json.JSONDecodeError as e:
            errors.append(f"{upload.filename}: JSON 파싱 실패"); logger.error(e)
        except Exception as e:
            errors.append(f"{upload.filename}: {e}"); logger.error(e)

    if results:
        await db.commit()
        # refresh all rows
        for r in results:
            result = await db.execute(select(Expense).where(Expense.id == r["id"]))
            row = result.scalar_one_or_none()
            if row:
                r.update(_to_dict(row))

    return {"results": results, "errors": errors, "count": len(results), "report_date": today}


@router.post("/export")
async def export_excel(req: ExportRequest, db: AsyncSession = Depends(get_db)):
    """DB 데이터 → Excel 다운로드"""
    result = await db.execute(
        select(Expense)
        .where(Expense.report_date == req.report_date)
        .order_by(Expense.id)
    )
    rows = result.scalars().all()
    if not rows:
        raise HTTPException(status_code=404, detail=f"{req.report_date} 데이터 없음")

    rows_dict = [_to_dict(r) for r in rows]
    excel     = _build_excel(req.report_date, rows_dict)
    filename  = f"Expense_Report_{req.report_date}.xlsx"
    return StreamingResponse(
        io.BytesIO(excel),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
