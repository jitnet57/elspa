"""
SSS Electronic Contribution — Image Scan → DB 저장 + Excel + CRUD
POST /api/sss/scan              : 이미지 업로드 → OCR → DB 저장 → Excel 반환
GET  /api/sss/months            : 저장된 월 목록
GET  /api/sss/records           : 특정 월 직원 목록 (?month=September+2024)
POST /api/sss/records           : 직원 1명 추가
PUT  /api/sss/records/{id}      : 직원 수정
DELETE /api/sss/records/{id}    : 직원 삭제
GET  /api/sss/export            : DB 데이터 → Excel 다운로드 (?month=...)
"""

import io
import base64
import json
import logging
from typing import List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, UploadFile, File, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete, func as sqlfunc
import anthropic

from app.config import settings
from app.database import get_db
from app.models.sss_contribution import SssContribution

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/sss", tags=["sss"])

ALLOWED_TYPES = {"image/jpeg", "image/jpg", "image/png", "image/webp", "image/gif"}

# ─── Pydantic 스키마 ────────────────────────────────────────────────

class SssRecordCreate(BaseModel):
    company:          Optional[str] = None
    employer_ss_no:   Optional[str] = None
    applicable_month: str
    invoice_no:       Optional[str] = None
    employee_no:      Optional[int] = None
    employee_name:    str = ""
    ss_number:        Optional[str] = None
    ss_amount:        float = 0
    ec_amount:        float = 0
    total_amount:     float = 0

class SssRecordUpdate(BaseModel):
    company:          Optional[str] = None
    employer_ss_no:   Optional[str] = None
    applicable_month: Optional[str] = None
    invoice_no:       Optional[str] = None
    employee_no:      Optional[int] = None
    employee_name:    Optional[str] = None
    ss_number:        Optional[str] = None
    ss_amount:        Optional[float] = None
    ec_amount:        Optional[float] = None
    total_amount:     Optional[float] = None

def _row_to_dict(row: SssContribution) -> dict:
    return {
        "id":               row.id,
        "company":          row.company,
        "employer_ss_no":   row.employer_ss_no,
        "applicable_month": row.applicable_month,
        "invoice_no":       row.invoice_no,
        "employee_no":      row.employee_no,
        "employee_name":    row.employee_name,
        "ss_number":        row.ss_number,
        "ss_amount":        float(row.ss_amount or 0),
        "ec_amount":        float(row.ec_amount or 0),
        "total_amount":     float(row.total_amount or 0),
        "created_at":       row.created_at.isoformat() if row.created_at else None,
        "updated_at":       row.updated_at.isoformat() if row.updated_at else None,
    }

# ─── OCR 프롬프트 ────────────────────────────────────────────────────

EXTRACT_PROMPT = """
You are an expert OCR assistant reading a Philippine SSS (Social Security System)
Electronic Contribution Collection List document.

Extract ALL data from this image and return ONLY valid JSON in this exact format:

{
  "doc_type": "summary" | "employee_list" | "unknown",
  "company": "company name if visible",
  "employer_ss_no": "employer SS number if visible",
  "applicable_month": "month and year if visible",
  "invoice_no": "invoice number if visible",
  "employees": [
    {
      "no": 1,
      "name": "LAST, FIRST MIDDLE",
      "ss_number": "XX-XXXXXXX-X",
      "ss": 1610.00,
      "ec": 10.00,
      "total": 1620.00
    }
  ],
  "totals": {
    "ss": 25410.00,
    "ec": 190.00,
    "total": 25600.00
  }
}

Rules:
- For "summary" pages (payment slip), set employees to []
- For "employee_list" pages, extract every row including rows with 0.00 values
- All monetary values must be numbers (not strings)
- If a field is not visible, use null
- Return ONLY the JSON object, no markdown, no explanation
"""

async def ocr_image(image_bytes: bytes, media_type: str) -> dict:
    """Claude Vision API로 SSS 이미지 OCR"""
    client = anthropic.Anthropic(api_key=settings.anthropic_api_key)
    b64 = base64.standard_b64encode(image_bytes).decode("utf-8")

    response = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=4096,
        messages=[{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
                {"type": "text", "text": EXTRACT_PROMPT},
            ],
        }],
    )

    raw = response.content[0].text.strip()
    if "```" in raw:
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    return json.loads(raw)

# ─── Excel 빌더 ─────────────────────────────────────────────────────

def build_excel(rows: list[dict], meta: dict) -> bytes:
    PESO   = '#,##0.00'
    thin   = Side(style="thin",   color="CCCCCC")
    med    = Side(style="medium", color="666666")
    B_THIN = Border(left=thin, right=thin, top=thin, bottom=thin)
    B_MED  = Border(left=med,  right=med,  top=med,  bottom=med)
    B_TOP  = Border(left=thin, right=thin, top=med,  bottom=med)

    def st(ws, r, c, val=None, bold=False, size=10, color="000000",
           bg=None, ha="left", va="center", bdr=B_THIN, fmt=None, wrap=False):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font      = Font(name="Calibri", bold=bold, size=size, color=color)
        cell.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
        if bg:  cell.fill = PatternFill("solid", fgColor=bg)
        if bdr: cell.border = bdr
        if fmt: cell.number_format = fmt
        return cell

    wb = openpyxl.Workbook()

    # ══ Sheet 1: Employee List ═══════════════════════════════════
    ws = wb.active
    ws.title = meta.get("applicable_month", "SSS Contribution")
    for i, w in enumerate([5, 32, 18, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:F1")
    st(ws, 1, 1, "SOCIAL SECURITY SYSTEM — Electronic Contribution Collection List",
       bold=True, size=13, color="FFFFFF", bg="1F3864", ha="center", bdr=B_MED)
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:F2")
    st(ws, 2, 1, meta.get("company", "—"), size=10, bg="BDD7EE", ha="center")

    pairs = [
        ("Employer SS No:", meta.get("employer_ss_no", "—")),
        ("Month:",          meta.get("applicable_month", "—")),
        ("Invoice:",        meta.get("invoice_no", "—")),
    ]
    for ci, (lbl, val) in zip([1, 3, 5], pairs):
        st(ws, 3, ci,   lbl, bold=True, size=9, bg="FFF2CC", ha="right")
        st(ws, 3, ci+1, val, size=9,    bg="FFF2CC", ha="left")
    ws.row_dimensions[4].height = 5

    hdrs = ["No.", "Name of Employee", "SS Number", "SS (PHP)", "EC (PHP)", "Total (PHP)"]
    for ci, h in enumerate(hdrs, 1):
        st(ws, 5, ci, h, bold=True, size=10, color="FFFFFF",
           bg="2E75B6", ha="center", bdr=B_MED, wrap=True)
    ws.row_dimensions[5].height = 26

    for idx, emp in enumerate(rows):
        row = idx + 6
        bg  = "F5F5F5" if idx % 2 == 0 else "FFFFFF"
        ss_v  = float(emp.get("ss_amount", 0) or 0)
        ec_v  = float(emp.get("ec_amount", 0) or 0)
        tot_v = float(emp.get("total_amount", 0) or 0)
        st(ws, row, 1, emp.get("employee_no"),           size=9, ha="center", bg=bg)
        st(ws, row, 2, emp.get("employee_name", ""),     size=9, ha="left",   bg=bg)
        st(ws, row, 3, emp.get("ss_number", ""),         size=9, ha="center", bg=bg)
        st(ws, row, 4, ss_v,  size=9, ha="right", bg="FCE4D6" if ss_v  == 0 else bg, fmt=PESO)
        st(ws, row, 5, ec_v,  size=9, ha="right", bg="FCE4D6" if ec_v  == 0 else bg, fmt=PESO)
        st(ws, row, 6, tot_v, size=9, ha="right", bg="FCE4D6" if tot_v == 0 else bg, fmt=PESO)
        ws.row_dimensions[row].height = 15

    tr = 6 + len(rows)
    ws.merge_cells(f"A{tr}:C{tr}")
    st(ws, tr, 1, "TOTAL", bold=True, size=11, color="FFFFFF", bg="2E75B6", ha="center", bdr=B_TOP)
    ss_t  = sum(float(r.get("ss_amount",    0) or 0) for r in rows)
    ec_t  = sum(float(r.get("ec_amount",    0) or 0) for r in rows)
    tot_t = sum(float(r.get("total_amount", 0) or 0) for r in rows)
    for ci, val in zip([4, 5, 6], [ss_t, ec_t, tot_t]):
        st(ws, tr, ci, val, bold=True, size=11, bg="E2EFDA", ha="right", fmt=PESO, bdr=B_TOP)
    ws.row_dimensions[tr].height = 20
    ws.freeze_panes = "A6"

    # ══ Sheet 2: Summary ═════════════════════════════════════════
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 24
    ws2.merge_cells("A1:B1")
    st(ws2, 1, 1, "SSS PAYMENT SUMMARY", bold=True, size=14,
       color="FFFFFF", bg="1F3864", ha="center", bdr=B_MED)
    ws2.row_dimensions[1].height = 24

    for ri, (lbl, val) in enumerate([
        ("Company",           meta.get("company", "—")),
        ("Employer SS No.",   meta.get("employer_ss_no", "—")),
        ("Applicable Month",  meta.get("applicable_month", "—")),
        ("Invoice No.",       meta.get("invoice_no", "—")),
        ("", ""),
        ("Total Employees",   len(rows)),
        ("Employees w/ SS>0", sum(1 for r in rows if float(r.get("ss_amount", 0) or 0) > 0)),
        ("", ""),
        ("SS Contribution",   ss_t),
        ("EC Contribution",   ec_t),
        ("TOTAL AMOUNT DUE",  tot_t),
    ], 2):
        is_total = lbl == "TOTAL AMOUNT DUE"
        fmt = PESO if isinstance(val, float) else None
        st(ws2, ri, 1, lbl, bold=is_total, size=10,
           bg="F4B942" if is_total else "BDD7EE", ha="left")
        st(ws2, ri, 2, val, bold=is_total, size=10,
           bg="FFF2CC" if is_total else "FFFFFF", ha="right", fmt=fmt)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ─── 엔드포인트 ─────────────────────────────────────────────────────

@router.post("/scan")
async def scan_sss_images(
    files: List[UploadFile] = File(...),
    db: AsyncSession = Depends(get_db),
):
    """SSS 영수증 이미지 OCR → DB 저장 → Excel 반환"""
    if not files:
        raise HTTPException(status_code=400, detail="파일을 선택해주세요")

    for f in files:
        if f.content_type not in ALLOWED_TYPES:
            raise HTTPException(
                status_code=400,
                detail=f"지원하지 않는 파일: {f.filename} ({f.content_type})"
            )

    pages, errors = [], []
    for upload in files:
        try:
            content = await upload.read()
            data    = await ocr_image(content, upload.content_type)
            pages.append(data)
            logger.info(f"✅ OCR: {upload.filename} | {len(data.get('employees', []))}명")
        except json.JSONDecodeError as e:
            errors.append(f"{upload.filename}: JSON 파싱 실패")
            logger.error(e)
        except Exception as e:
            errors.append(f"{upload.filename}: {str(e)}")
            logger.error(e)

    if not pages:
        raise HTTPException(status_code=422, detail=f"OCR 실패: {'; '.join(errors)}")

    # 페이지 병합
    meta: dict = {}
    all_employees: list = []
    for page in pages:
        if not meta.get("company") and page.get("company"):
            meta.update({
                "company":          page.get("company"),
                "employer_ss_no":   page.get("employer_ss_no"),
                "applicable_month": page.get("applicable_month"),
                "invoice_no":       page.get("invoice_no"),
            })
        existing_names = {e["employee_name"] for e in all_employees}
        for emp in page.get("employees", []):
            if emp.get("name") and emp["name"] not in existing_names:
                all_employees.append({
                    "employee_name": emp["name"],
                    "ss_number":     emp.get("ss_number"),
                    "ss_amount":     float(emp.get("ss", 0) or 0),
                    "ec_amount":     float(emp.get("ec", 0) or 0),
                    "total_amount":  float(emp.get("total", 0) or 0),
                })
                existing_names.add(emp["name"])

    for i, emp in enumerate(all_employees, 1):
        emp["employee_no"] = i

    # DB 저장 — 같은 invoice_no 또는 (같은 month+company) 기존 데이터 교체
    month = meta.get("applicable_month")
    if month:
        await db.execute(
            delete(SssContribution).where(SssContribution.applicable_month == month)
        )
        for emp in all_employees:
            db.add(SssContribution(
                company          = meta.get("company"),
                employer_ss_no   = meta.get("employer_ss_no"),
                applicable_month = month,
                invoice_no       = meta.get("invoice_no"),
                **emp,
            ))
        await db.commit()
        logger.info(f"💾 DB 저장: {month} — {len(all_employees)}명")

    # Excel 생성
    rows_for_excel = [{**emp, **meta} for emp in all_employees]
    excel_bytes = build_excel(rows_for_excel, meta)
    month_slug  = (month or "Contribution").replace(" ", "_")
    filename    = f"SSS_{month_slug}.xlsx"

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/months")
async def get_months(db: AsyncSession = Depends(get_db)):
    """저장된 SSS 월 목록 + 요약"""
    result = await db.execute(
        select(
            SssContribution.applicable_month,
            SssContribution.company,
            SssContribution.employer_ss_no,
            SssContribution.invoice_no,
            sqlfunc.count(SssContribution.id).label("count"),
            sqlfunc.sum(SssContribution.total_amount).label("total"),
        )
        .group_by(
            SssContribution.applicable_month,
            SssContribution.company,
            SssContribution.employer_ss_no,
            SssContribution.invoice_no,
        )
        .order_by(SssContribution.applicable_month.desc())
    )
    rows = result.all()
    return [
        {
            "applicable_month": r.applicable_month,
            "company":          r.company,
            "employer_ss_no":   r.employer_ss_no,
            "invoice_no":       r.invoice_no,
            "count":            r.count,
            "total":            float(r.total or 0),
        }
        for r in rows
    ]


@router.get("/records")
async def get_records(
    month: str = Query(..., description="예: September 2024"),
    db: AsyncSession = Depends(get_db),
):
    """특정 월의 SSS 직원 목록"""
    result = await db.execute(
        select(SssContribution)
        .where(SssContribution.applicable_month == month)
        .order_by(SssContribution.employee_no)
    )
    rows = result.scalars().all()
    return [_row_to_dict(r) for r in rows]


@router.post("/records")
async def create_record(
    data: SssRecordCreate,
    db: AsyncSession = Depends(get_db),
):
    """직원 레코드 1건 추가"""
    row = SssContribution(**data.model_dump())
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return _row_to_dict(row)


@router.put("/records/{record_id}")
async def update_record(
    record_id: int,
    data: SssRecordUpdate,
    db: AsyncSession = Depends(get_db),
):
    """직원 레코드 수정"""
    result = await db.execute(
        select(SssContribution).where(SssContribution.id == record_id)
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="레코드를 찾을 수 없습니다")

    for field, value in data.model_dump(exclude_none=True).items():
        setattr(row, field, value)

    await db.commit()
    await db.refresh(row)
    return _row_to_dict(row)


@router.delete("/records/{record_id}")
async def delete_record(
    record_id: int,
    db: AsyncSession = Depends(get_db),
):
    """직원 레코드 삭제"""
    result = await db.execute(
        select(SssContribution).where(SssContribution.id == record_id)
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="레코드를 찾을 수 없습니다")

    await db.delete(row)
    await db.commit()
    return {"success": True, "deleted_id": record_id}


@router.get("/export")
async def export_month(
    month: str = Query(..., description="예: September 2024"),
    db: AsyncSession = Depends(get_db),
):
    """DB 데이터 → Excel 다운로드"""
    result = await db.execute(
        select(SssContribution)
        .where(SssContribution.applicable_month == month)
        .order_by(SssContribution.employee_no)
    )
    rows = result.scalars().all()
    if not rows:
        raise HTTPException(status_code=404, detail=f"'{month}' 데이터가 없습니다")

    rows_dict = [_row_to_dict(r) for r in rows]
    meta = {
        "company":          rows[0].company,
        "employer_ss_no":   rows[0].employer_ss_no,
        "applicable_month": month,
        "invoice_no":       rows[0].invoice_no,
    }
    excel_bytes = build_excel(rows_dict, meta)
    filename    = f"SSS_{month.replace(' ', '_')}.xlsx"

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
