"""
📌 Excel Import Service — 엑셀 파일 대량 데이터 임포트 시스템
📋 목적: 엑셀 파일 파싱, 검증, 변환, DB 저장 (트랜잭션 지원)
🔧 포함: 파일 파싱, 매핑 검증, 행 검증, 데이터 변환, DB 저장
📅 작성일: 2026-06-02

핵심 기능:
  1. parse_excel_file(file_path) - 엑셀 파일 파싱 및 샘플 데이터 추출
  2. validate_mapping(table_name, mapping) - 매핑 정의 검증
  3. validate_row(row, mapping, db_session) - 단일 행 검증
  4. transform_row_data(row, mapping) - 행 데이터 변환 (타입, 포맷)
  5. insert_rows_to_db(table_name, rows, mapping) - 트랜잭션 기반 DB 저장

사용 예:
  # 1단계: 파일 파싱
  result = parse_excel_file('/path/to/file.xlsx')

  # 2단계: 매핑 정의 및 검증
  mapping = {
      'therapist_id': {'column': 'Therapist ID', 'type': 'int', 'required': True},
      'name': {'column': 'Name', 'type': 'str', 'required': True},
      'salary': {'column': 'Salary', 'type': 'decimal', 'required': False}
  }
  validate_mapping('therapist', mapping)

  # 3단계: DB 저장 (트랜잭션)
  import_result = insert_rows_to_db(
      'therapist',
      result['sample_data'],
      mapping,
      progress_callback=lambda current, total: print(f"{current}/{total}")
  )
"""

import os
from typing import Dict, List, Any, Optional, Callable, Tuple
from decimal import Decimal
from datetime import datetime, date
from dataclasses import dataclass, field
import logging
from enum import Enum

from sqlalchemy.orm import Session
from sqlalchemy import inspect, text
from sqlalchemy.exc import SQLAlchemyError, IntegrityError

logger = logging.getLogger(__name__)


# ============================================================
# 에러 클래스 정의
# ============================================================

class ImportError(Exception):
    """엑셀 임포트 기본 에러"""
    pass


class FileParseError(ImportError):
    """파일 파싱 에러"""
    pass


class MappingValidationError(ImportError):
    """매핑 검증 에러"""
    pass


class RowValidationError(ImportError):
    """행 검증 에러"""
    pass


class DataTransformationError(ImportError):
    """데이터 변환 에러"""
    pass


class DatabaseInsertError(ImportError):
    """데이터베이스 저장 에러"""
    pass


# ============================================================
# 데이터 클래스
# ============================================================

@dataclass
class FieldError:
    """필드별 검증 에러"""
    field_name: str
    error_message: str
    row_number: int = 0
    suggested_value: Optional[str] = None


@dataclass
class RowError:
    """행별 검증 에러"""
    row_number: int
    field_errors: List[FieldError] = field(default_factory=list)
    general_error: Optional[str] = None

    def add_field_error(self, field_name: str, error_message: str, suggested_value: Optional[str] = None):
        """필드 에러 추가"""
        self.field_errors.append(
            FieldError(field_name, error_message, self.row_number, suggested_value)
        )

    def has_errors(self) -> bool:
        """에러 존재 여부"""
        return len(self.field_errors) > 0 or self.general_error is not None


@dataclass
class ImportResult:
    """임포트 결과"""
    success: bool
    total_rows: int
    inserted_rows: int
    skipped_rows: int
    failed_rows: int
    row_errors: List[RowError] = field(default_factory=list)
    general_errors: List[str] = field(default_factory=list)
    message: str = ""

    def add_row_error(self, row_error: RowError):
        """행 에러 추가"""
        self.row_errors.append(row_error)

    def add_general_error(self, error: str):
        """일반 에러 추가"""
        self.general_errors.append(error)


@dataclass
class ParseResult:
    """파일 파싱 결과"""
    headers: List[str]
    sample_data: List[Dict[str, Any]]
    total_rows: int
    errors: List[str] = field(default_factory=list)


@dataclass
class MappingValidationResult:
    """매핑 검증 결과"""
    is_valid: bool
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


# ============================================================
# 매핑 설정 타입
# ============================================================

class FieldType(str, Enum):
    """필드 타입"""
    STRING = "str"
    INTEGER = "int"
    DECIMAL = "decimal"
    FLOAT = "float"
    BOOLEAN = "bool"
    DATE = "date"
    DATETIME = "datetime"
    EMAIL = "email"


# ============================================================
# 핵심 함수 1: 파일 파싱
# ============================================================

def parse_excel_file(file_path: str, sheet_name: int = 0, max_sample_rows: int = 10) -> ParseResult:
    """
    엑셀 파일 파싱 및 메타데이터 추출

    Args:
        file_path: 엑셀 파일 경로
        sheet_name: 시트 인덱스 (기본값: 0)
        max_sample_rows: 샘플 데이터 최대 행 수

    Returns:
        ParseResult: {headers, sample_data, total_rows, errors}

    예외:
        FileParseError: 파일 읽기 실패 시

    구현 예:
        >>> result = parse_excel_file('/path/to/employees.xlsx')
        >>> print(result.headers)  # ['Name', 'Email', 'Salary']
        >>> print(result.total_rows)  # 100
        >>> print(len(result.sample_data))  # 10 (max_sample_rows)
    """
    try:
        # openpyxl 임포트 시도
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise FileParseError(
                "openpyxl 라이브러리가 설치되지 않았습니다. "
                "pip install openpyxl 를 실행해주세요."
            )

        # 파일 존재 여부 확인
        if not os.path.exists(file_path):
            raise FileParseError(f"파일을 찾을 수 없습니다: {file_path}")

        # 파일 형식 확인
        if not file_path.lower().endswith(('.xlsx', '.xls')):
            raise FileParseError(f"지원하지 않는 파일 형식입니다: {file_path}")

        # 워크북 로드
        workbook = load_workbook(file_path, data_only=True)

        # 시트 가져오기
        if isinstance(sheet_name, int):
            if sheet_name >= len(workbook.sheetnames):
                raise FileParseError(
                    f"시트 인덱스가 범위를 벗어났습니다. "
                    f"사용 가능한 시트: {workbook.sheetnames}"
                )
            worksheet = workbook.worksheets[sheet_name]
        else:
            if sheet_name not in workbook.sheetnames:
                raise FileParseError(f"시트를 찾을 수 없습니다: {sheet_name}")
            worksheet = workbook[sheet_name]

        # 헤더 추출 (첫 번째 행)
        headers = []
        for cell in worksheet[1]:
            if cell.value is not None:
                headers.append(str(cell.value).strip())

        if not headers:
            raise FileParseError("헤더를 찾을 수 없습니다. 첫 번째 행이 비어있습니다.")

        # 데이터 행 추출
        sample_data = []
        total_rows = 0

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), start=2):
            # 행의 모든 셀이 비어있으면 건너뛰기
            row_values = [cell.value for cell in row]
            if all(v is None for v in row_values):
                break

            total_rows += 1

            # 샘플 데이터 수집
            if len(sample_data) < max_sample_rows:
                row_dict = {}
                for header, cell in zip(headers, row):
                    row_dict[header] = cell.value
                sample_data.append(row_dict)

        workbook.close()

        return ParseResult(
            headers=headers,
            sample_data=sample_data,
            total_rows=total_rows
        )

    except FileParseError:
        raise
    except Exception as e:
        logger.exception(f"엑셀 파일 파싱 오류: {str(e)}")
        raise FileParseError(f"엑셀 파일 파싱 실패: {str(e)}")


# ============================================================
# 핵심 함수 2: 매핑 검증
# ============================================================

def validate_mapping(
    table_name: str,
    mapping: Dict[str, Dict[str, Any]],
    db_session: Optional[Session] = None
) -> MappingValidationResult:
    """
    매핑 정의 검증

    Args:
        table_name: 테이블명 (예: 'therapist', 'staff')
        mapping: 매핑 정의
            {
                'field_name': {
                    'column': '엑셀 컬럼명',
                    'type': 'str|int|decimal|date|datetime|boolean|email',
                    'required': True|False,
                    'unique': True|False (선택사항),
                    'validator': callable (선택사항),
                    'default': 기본값 (선택사항)
                },
                ...
            }
        db_session: DB 세션 (선택사항, FK 검증용)

    Returns:
        MappingValidationResult: {is_valid, errors, warnings}

    예외:
        MappingValidationError: 매핑 정의 오류 시

    구현 예:
        >>> mapping = {
        ...     'therapist_id': {'column': 'ID', 'type': 'int', 'required': True},
        ...     'name': {'column': 'Name', 'type': 'str', 'required': True},
        ...     'salary': {'column': 'Salary', 'type': 'decimal', 'required': False}
        ... }
        >>> result = validate_mapping('therapist', mapping)
        >>> assert result.is_valid
    """
    errors = []
    warnings = []

    # 매핑이 비어있는지 확인
    if not mapping:
        errors.append("매핑 정의가 비어있습니다.")
        return MappingValidationResult(is_valid=False, errors=errors)

    # 각 필드 검증
    for field_name, field_config in mapping.items():
        # 필수 키 확인
        if not isinstance(field_config, dict):
            errors.append(f"필드 '{field_name}'의 설정이 딕셔너리가 아닙니다.")
            continue

        if 'column' not in field_config:
            errors.append(f"필드 '{field_name}'에 'column' 키가 없습니다.")

        if 'type' not in field_config:
            errors.append(f"필드 '{field_name}'에 'type' 키가 없습니다.")
        else:
            # 타입 검증
            field_type = field_config['type']
            if field_type not in [ft.value for ft in FieldType]:
                errors.append(
                    f"필드 '{field_name}'의 타입이 유효하지 않습니다: {field_type}. "
                    f"사용 가능한 타입: {', '.join([ft.value for ft in FieldType])}"
                )

        # 선택사항 키 검증
        if 'validator' in field_config:
            if not callable(field_config['validator']):
                errors.append(f"필드 '{field_name}'의 validator가 callable이 아닙니다.")

        # 유니크 제약 검증
        if field_config.get('unique', False) and not field_config.get('required', False):
            warnings.append(
                f"필드 '{field_name}'이 유니크하지만 필수가 아닙니다. "
                "NULL 값으로 인한 유니크 제약 위반이 발생할 수 있습니다."
            )

    is_valid = len(errors) == 0

    return MappingValidationResult(
        is_valid=is_valid,
        errors=errors,
        warnings=warnings
    )


# ============================================================
# 핵심 함수 3: 행 검증
# ============================================================

def validate_row(
    row: Dict[str, Any],
    mapping: Dict[str, Dict[str, Any]],
    db_session: Optional[Session] = None,
    row_number: int = 0
) -> RowError:
    """
    단일 행의 데이터 검증

    Args:
        row: 행 데이터 딕셔너리
        mapping: 매핑 정의
        db_session: DB 세션 (FK 검증용)
        row_number: 행 번호 (에러 메시지용)

    Returns:
        RowError: 검증 결과 및 에러 정보

    예외:
        RowValidationError: 치명적 검증 오류 시

    구현 예:
        >>> row = {'Name': 'John', 'Email': 'john@example.com', 'Salary': '50000'}
        >>> mapping = {
        ...     'name': {'column': 'Name', 'type': 'str', 'required': True},
        ...     'email': {'column': 'Email', 'type': 'email', 'required': True},
        ...     'salary': {'column': 'Salary', 'type': 'decimal', 'required': False}
        ... }
        >>> row_error = validate_row(row, mapping, row_number=2)
        >>> print(row_error.has_errors())  # False if valid
    """
    row_error = RowError(row_number=row_number)

    for field_name, field_config in mapping.items():
        column_name = field_config.get('column')
        field_type = field_config.get('type')
        is_required = field_config.get('required', False)
        validator = field_config.get('validator')

        # 컬럼값 추출
        cell_value = row.get(column_name)

        # 필수 필드 검증
        if is_required and (cell_value is None or str(cell_value).strip() == ''):
            row_error.add_field_error(
                field_name,
                f"필수 필드가 비어있습니다."
            )
            continue

        # 비어있는 경우 건너뛰기
        if cell_value is None or str(cell_value).strip() == '':
            continue

        # 타입별 검증
        try:
            if field_type == FieldType.INTEGER:
                try:
                    int(str(cell_value).strip())
                except ValueError:
                    row_error.add_field_error(
                        field_name,
                        f"정수 형식이 아닙니다: '{cell_value}'",
                        suggested_value=None
                    )

            elif field_type == FieldType.DECIMAL:
                try:
                    Decimal(str(cell_value).strip())
                except:
                    row_error.add_field_error(
                        field_name,
                        f"숫자 형식이 아닙니다: '{cell_value}'",
                        suggested_value=None
                    )

            elif field_type == FieldType.FLOAT:
                try:
                    float(str(cell_value).strip())
                except ValueError:
                    row_error.add_field_error(
                        field_name,
                        f"실수 형식이 아닙니다: '{cell_value}'",
                        suggested_value=None
                    )

            elif field_type == FieldType.EMAIL:
                # 간단한 이메일 검증
                email_str = str(cell_value).strip()
                if '@' not in email_str or '.' not in email_str:
                    row_error.add_field_error(
                        field_name,
                        f"유효한 이메일 주소가 아닙니다: '{email_str}'",
                        suggested_value=None
                    )

            elif field_type == FieldType.DATE:
                # 날짜 형식 검증
                try:
                    if isinstance(cell_value, date):
                        pass
                    else:
                        datetime.strptime(str(cell_value).strip(), "%Y-%m-%d")
                except ValueError:
                    row_error.add_field_error(
                        field_name,
                        f"날짜 형식이 아닙니다: '{cell_value}' (예: 2025-06-02)",
                        suggested_value=None
                    )

            elif field_type == FieldType.DATETIME:
                # 날짜시간 형식 검증
                try:
                    if isinstance(cell_value, datetime):
                        pass
                    else:
                        datetime.fromisoformat(str(cell_value).strip())
                except ValueError:
                    row_error.add_field_error(
                        field_name,
                        f"날짜시간 형식이 아닙니다: '{cell_value}' (예: 2025-06-02T14:30:00)",
                        suggested_value=None
                    )

            elif field_type == FieldType.BOOLEAN:
                # 불린 형식 검증
                bool_str = str(cell_value).strip().lower()
                if bool_str not in ['true', 'false', '1', '0', 'yes', 'no', 'y', 'n']:
                    row_error.add_field_error(
                        field_name,
                        f"불린 형식이 아닙니다: '{cell_value}' (예: true, false, yes, no)",
                        suggested_value=None
                    )

        except Exception as e:
            row_error.add_field_error(
                field_name,
                f"타입 검증 중 오류: {str(e)}",
                suggested_value=None
            )

        # 커스텀 검증 함수
        if validator and callable(validator):
            try:
                is_valid, error_message = validator(cell_value)
                if not is_valid:
                    row_error.add_field_error(
                        field_name,
                        error_message,
                        suggested_value=None
                    )
            except Exception as e:
                row_error.add_field_error(
                    field_name,
                    f"커스텀 검증 중 오류: {str(e)}",
                    suggested_value=None
                )

    return row_error


# ============================================================
# 핵심 함수 4: 데이터 변환
# ============================================================

def transform_row_data(
    row: Dict[str, Any],
    mapping: Dict[str, Dict[str, Any]]
) -> Dict[str, Any]:
    """
    행 데이터를 DB 저장용으로 변환

    Args:
        row: 원본 행 데이터
        mapping: 매핑 정의

    Returns:
        Dict[str, Any]: 변환된 데이터
            {
                'field_name': 변환된_값,
                ...
            }

    예외:
        DataTransformationError: 변환 실패 시

    구현 예:
        >>> row = {'Name': 'John', 'Salary': '50000.50', 'Active': 'yes'}
        >>> mapping = {
        ...     'name': {'column': 'Name', 'type': 'str'},
        ...     'salary': {'column': 'Salary', 'type': 'decimal'},
        ...     'active': {'column': 'Active', 'type': 'boolean'}
        ... }
        >>> transformed = transform_row_data(row, mapping)
        >>> print(transformed)  # {'name': 'John', 'salary': Decimal('50000.50'), 'active': True}
    """
    transformed = {}

    for field_name, field_config in mapping.items():
        column_name = field_config.get('column')
        field_type = field_config.get('type')
        default_value = field_config.get('default')

        # 컬럼값 추출
        cell_value = row.get(column_name)

        # NULL 값 처리
        if cell_value is None or str(cell_value).strip() == '':
            if default_value is not None:
                transformed[field_name] = default_value
            else:
                transformed[field_name] = None
            continue

        try:
            # 타입별 변환
            if field_type == FieldType.STRING:
                transformed[field_name] = str(cell_value).strip()

            elif field_type == FieldType.INTEGER:
                transformed[field_name] = int(str(cell_value).strip())

            elif field_type == FieldType.DECIMAL:
                transformed[field_name] = Decimal(str(cell_value).strip())

            elif field_type == FieldType.FLOAT:
                transformed[field_name] = float(str(cell_value).strip())

            elif field_type == FieldType.BOOLEAN:
                bool_str = str(cell_value).strip().lower()
                transformed[field_name] = bool_str in ['true', '1', 'yes', 'y']

            elif field_type == FieldType.DATE:
                if isinstance(cell_value, date):
                    transformed[field_name] = cell_value
                else:
                    transformed[field_name] = datetime.strptime(
                        str(cell_value).strip(), "%Y-%m-%d"
                    ).date()

            elif field_type == FieldType.DATETIME:
                if isinstance(cell_value, datetime):
                    transformed[field_name] = cell_value
                else:
                    transformed[field_name] = datetime.fromisoformat(
                        str(cell_value).strip()
                    )

            elif field_type == FieldType.EMAIL:
                transformed[field_name] = str(cell_value).strip().lower()

            else:
                transformed[field_name] = cell_value

        except Exception as e:
            raise DataTransformationError(
                f"필드 '{field_name}' 변환 실패: {str(e)}"
            )

    return transformed


# ============================================================
# 핵심 함수 5: DB에 저장
# ============================================================

def insert_rows_to_db(
    table_name: str,
    rows: List[Dict[str, Any]],
    mapping: Dict[str, Dict[str, Any]],
    db_session: Session,
    model_class: Any,
    progress_callback: Optional[Callable[[int, int], None]] = None,
    skip_errors: bool = False
) -> ImportResult:
    """
    행 데이터를 데이터베이스에 저장 (트랜잭션 지원)

    Args:
        table_name: 테이블명
        rows: 행 데이터 리스트
        mapping: 매핑 정의
        db_session: SQLAlchemy 세션
        model_class: SQLAlchemy 모델 클래스
        progress_callback: 진행률 콜백 함수 (current, total)
        skip_errors: True면 에러 발생 시 행 스킵, False면 트랜잭션 롤백

    Returns:
        ImportResult: {
            'success': bool,
            'total_rows': int,
            'inserted_rows': int,
            'skipped_rows': int,
            'failed_rows': int,
            'row_errors': List[RowError],
            'general_errors': List[str],
            'message': str
        }

    예외:
        DatabaseInsertError: DB 오류 시 (skip_errors=False)

    구현 예:
        >>> from app.models.therapist import Therapist
        >>> rows = [
        ...     {'Name': 'John', 'Email': 'john@example.com'},
        ...     {'Name': 'Jane', 'Email': 'jane@example.com'}
        ... ]
        >>> mapping = {
        ...     'name': {'column': 'Name', 'type': 'str', 'required': True},
        ...     'email': {'column': 'Email', 'type': 'email', 'required': True}
        ... }
        >>> result = insert_rows_to_db(
        ...     'therapist', rows, mapping, db_session, Therapist,
        ...     progress_callback=lambda c, t: print(f"{c}/{t}")
        ... )
        >>> print(f"성공: {result.inserted_rows}, 실패: {result.failed_rows}")
    """
    import_result = ImportResult(
        success=False,
        total_rows=len(rows),
        inserted_rows=0,
        skipped_rows=0,
        failed_rows=0
    )

    if not rows:
        import_result.message = "임포트할 행이 없습니다."
        import_result.success = True
        return import_result

    # 트랜잭션 시작
    try:
        for row_idx, row in enumerate(rows, start=2):  # 헤더를 1로 계산
            # 진행률 콜백
            if progress_callback:
                progress_callback(row_idx - 1, len(rows))

            # 행 검증
            row_error = validate_row(row, mapping, db_session, row_idx)

            if row_error.has_errors():
                import_result.skipped_rows += 1
                import_result.add_row_error(row_error)

                if not skip_errors:
                    import_result.failed_rows += 1
                    import_result.add_general_error(
                        f"행 {row_idx}에서 검증 오류 발생: "
                        f"{len(row_error.field_errors)}개 필드 오류"
                    )
                    # skip_errors=False면 여기서 중단하지 않고, 마지막에 롤백
                continue

            # 데이터 변환
            try:
                transformed_data = transform_row_data(row, mapping)
            except DataTransformationError as e:
                import_result.skipped_rows += 1
                row_error.general_error = str(e)
                import_result.add_row_error(row_error)

                if not skip_errors:
                    import_result.failed_rows += 1
                continue

            # DB에 저장
            try:
                instance = model_class(**transformed_data)
                db_session.add(instance)
                db_session.flush()  # 에러 즉시 감지
                import_result.inserted_rows += 1

            except IntegrityError as e:
                db_session.rollback()
                import_result.skipped_rows += 1
                row_error.general_error = f"무결성 제약 위반: {str(e.orig)}"
                import_result.add_row_error(row_error)

                if not skip_errors:
                    import_result.failed_rows += 1
                else:
                    db_session.rollback()  # flush된 데이터 취소

            except SQLAlchemyError as e:
                db_session.rollback()
                import_result.skipped_rows += 1
                row_error.general_error = f"DB 오류: {str(e)}"
                import_result.add_row_error(row_error)

                if not skip_errors:
                    import_result.failed_rows += 1
                else:
                    db_session.rollback()

        # 에러가 있으면 롤백
        if import_result.failed_rows > 0 and not skip_errors:
            db_session.rollback()
            import_result.success = False
            import_result.message = (
                f"검증 오류로 인해 임포트가 중단되었습니다. "
                f"실패한 행: {import_result.failed_rows}"
            )
            return import_result

        # 최종 커밋
        db_session.commit()
        import_result.success = True
        import_result.message = (
            f"임포트 완료: {import_result.inserted_rows}개 행 저장, "
            f"{import_result.skipped_rows}개 행 스킵"
        )

        logger.info(import_result.message)
        return import_result

    except Exception as e:
        db_session.rollback()
        logger.exception(f"DB 저장 중 오류: {str(e)}")
        import_result.success = False
        import_result.add_general_error(f"DB 저장 실패: {str(e)}")
        import_result.message = f"임포트 실패: {str(e)}"
        return import_result


# ============================================================
# 유틸리티 함수
# ============================================================

def create_default_mapping(
    model_class: Any,
    field_overrides: Optional[Dict[str, Dict[str, Any]]] = None
) -> Dict[str, Dict[str, Any]]:
    """
    SQLAlchemy 모델에서 기본 매핑 생성

    Args:
        model_class: SQLAlchemy 모델 클래스
        field_overrides: 필드별 오버라이드 설정

    Returns:
        Dict: 기본 매핑

    구현 예:
        >>> from app.models.therapist import Therapist
        >>> mapping = create_default_mapping(
        ...     Therapist,
        ...     field_overrides={
        ...         'email': {'type': 'email', 'required': True}
        ...     }
        ... )
    """
    mapper = inspect(model_class)
    mapping = {}

    for column in mapper.columns:
        field_name = column.name

        # 기본 타입 결정
        if column.type.python_type == int:
            field_type = FieldType.INTEGER
        elif column.type.python_type == str:
            field_type = FieldType.STRING
        elif column.type.python_type == Decimal:
            field_type = FieldType.DECIMAL
        elif column.type.python_type == float:
            field_type = FieldType.FLOAT
        elif column.type.python_type == bool:
            field_type = FieldType.BOOLEAN
        elif column.type.python_type == date:
            field_type = FieldType.DATE
        elif column.type.python_type == datetime:
            field_type = FieldType.DATETIME
        else:
            field_type = FieldType.STRING

        # 필드 설정
        mapping[field_name] = {
            'column': field_name.replace('_', ' ').title(),
            'type': field_type,
            'required': not column.nullable
        }

        # 오버라이드 적용
        if field_overrides and field_name in field_overrides:
            mapping[field_name].update(field_overrides[field_name])

    return mapping
