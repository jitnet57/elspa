"""
Daily Expense Report — Receipt Scan → Excel
POST /api/expense/scan   : 영수증 이미지 업로드 → Claude Vision OCR → JSON 반환
POST /api/expense/export : JSON 지출 데이터 → 포맷된 Excel 다운로드
"""

import io
import base64
import json
import logging
from typing import List, Optional
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import anthropic

from app.config import settings

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/expense", tags=["expense"])

ALLOWED_TYPES = {"image/jpeg", "image/jpg", "image/png", "image/webp", "image/gif"}

CATEGORIES = {
    "food":          "🍽️ Food (식비)",
    "transport":     "🚗 Transport (교통비)",
    "supplies":      "🛒 Supplies (소모품)",
    "utilities":     "💡 Utilities (공과금)",
    "entertainment": "🤝 Entertainment (접대비)",
    "medical":       "🏥 Medical (의료비)",
    "other":         "📦 Other (기타)",
}

# Excel 카테고리 배경색 (openpyxl용 hex)
CATEGORY_COLORS = {
    "food":          "FFD9B3",
    "transport":     "B3E5FC",
    "supplies":      "C8E6C9",
    "utilities":     "FFF9C4",
    "entertainment": "F8BBD9",
    "medical":       "E1BEE7",
    "other":         "EEEEEE",
}

EXTRACT_PROMPT = """
You are an expert OCR assistant reading a receipt or expense document (any language).

Extract the data and return ONLY valid JSON in this exact format:

{
  "vendor": "store or vendor name",
  "date": "date as shown on receipt, format YYYY-MM-DD if possible",
  "total_amount": 123.45,
  "currency": "PHP",
  "items": ["item description 1", "item description 2", "item description 3"],
  "category_guess": "food",
  "tax": 0.00,
  "notes": "receipt number, address, or other relevant info"
}

category_guess must be exactly one of:
- food: restaurants, cafes, groceries, food delivery, snacks
- transport: gas/fuel, taxi, Grab, parking, tolls, bus/train, ferry
- supplies: office supplies, cleaning materials, hardware, tools
- utilities: electricity, water, internet, phone bills, rent
- entertainment: team meals, client entertainment, events, gifts
- medical: medicine, clinic, hospital, health supplements
- other: anything that doesn't fit the above

Rules:
- total_amount must be a number (not a string). Use 0 if amount is not readable.
- items: brief descriptions, max 5 items
- If a field is not visible or unclear, use null
- Return ONLY the JSON object, no markdown, no explanation
"""


class ExpenseItem(BaseModel):
    id: int
    vendor: Optional[str] = None
    date: Optional[str] = None
    amount: float = 0
    currency: str = "PHP"
    category: str = "other"
    items: List[str] = []
    notes: Optional[str] = None
    filename: Optional[str] = None


class ExportRequest(BaseModel):
    report_date: str   # YYYY-MM-DD
    expenses: List[ExpenseItem]


# ─── OCR ──────────────────────────────────────────────────────────────────────

async def ocr_receipt(image_bytes: bytes, media_type: str, filename: str) -> dict:
    """Claude Vision API로 영수증에서 지출 데이터 추출"""
    client = anthropic.Anthropic(api_key=settings.anthropic_api_key)
    b64 = base64.standard_b64encode(image_bytes).decode("utf-8")

    # HEIC → JPEG로 처리
    if media_type in ("image/heic", "image/heif"):
        media_type = "image/jpeg"

    response = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=1024,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {"type": "base64", "media_type": media_type, "data": b64},
                },
                {"type": "text", "text": EXTRACT_PROMPT},
            ],
        }],
    )

    raw = response.content[0].text.strip()
    if "```" in raw:
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]

    data = json.loads(raw)
    data["filename"] = filename
    return data


# ─── Excel 생성 ───────────────────────────────────────────────────────────────

def build_expense_excel(report_date: str, expenses: list) -> bytes:
    """지출 데이터로 포맷된 Excel 파일 생성 (2 시트)"""
    PESO = '#,##0.00'
    thin = Side(style="thin", color="CCCCCC")
    med  = Side(style="medium", color="555555")
    B_THIN = Border(left=thin, right=thin, top=thin, bottom=thin)
    B_MED  = Border(left=med,  right=med,  top=med,  bottom=med)
    B_TOP  = Border(left=thin, right=thin, top=med,  bottom=med)

    def st(ws, r, c, val=None, bold=False, size=10, color="000000",
           bg=None, ha="left", va="center", bdr=B_THIN, fmt=None, wrap=False):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font      = Font(name="Calibri", bold=bold, size=size, color=color)
        cell.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        if bdr:
            cell.border = bdr
        if fmt:
            cell.number_format = fmt
        return cell

    wb = openpyxl.Workbook()

    # ══ Sheet 1: Daily Expenses ═══════════════════════════════════════════════
    ws = wb.active
    ws.title = f"Daily {report_date}"

    col_widths = [5, 22, 12, 26, 14, 34, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 타이틀
    ws.merge_cells("A1:G1")
    st(ws, 1, 1, f"DAILY EXPENSE REPORT — {report_date}",
       bold=True, size=13, color="FFFFFF", bg="1B5E20", ha="center", bdr=B_MED)
    ws.row_dimensions[1].height = 22

    # 생성 일시
    ws.merge_cells("A2:G2")
    st(ws, 2, 1,
       f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  ElSpa Management System",
       size=9, bg="C8E6C9", ha="center")
    ws.row_dimensions[2].height = 14

    ws.row_dimensions[3].height = 5  # 공백

    # 컬럼 헤더
    hdrs = ["No.", "Vendor", "Date", "Category", "Amount (PHP)", "Items", "Notes"]
    for ci, h in enumerate(hdrs, 1):
        st(ws, 4, ci, h, bold=True, size=10, color="FFFFFF",
           bg="2E7D32", ha="center", bdr=B_MED, wrap=True)
    ws.row_dimensions[4].height = 22

    # 데이터 행
    grand_total = 0
    for idx, exp in enumerate(expenses):
        row = idx + 5
        row_bg = "F1F8E9" if idx % 2 == 0 else "FFFFFF"
        amt = exp.get("amount", 0) or 0
        grand_total += amt
        cat_key   = exp.get("category", "other") or "other"
        cat_label = CATEGORIES.get(cat_key, "📦 Other")
        cat_bg    = CATEGORY_COLORS.get(cat_key, "EEEEEE")
        items_str = ", ".join(exp.get("items") or [])

        st(ws, row, 1, idx + 1,                    size=9, ha="center", bg=row_bg)
        st(ws, row, 2, exp.get("vendor") or "—",   size=9, ha="left",   bg=row_bg)
        st(ws, row, 3, exp.get("date") or "—",     size=9, ha="center", bg=row_bg)
        st(ws, row, 4, cat_label,                  size=9, ha="left",   bg=cat_bg)
        st(ws, row, 5, amt,                        size=9, ha="right",  bg=row_bg, fmt=PESO)
        st(ws, row, 6, items_str,                  size=9, ha="left",   bg=row_bg, wrap=True)
        st(ws, row, 7, exp.get("notes") or "",     size=9, ha="left",   bg=row_bg)
        ws.row_dimensions[row].height = 15

    # 합계 행
    tr = 5 + len(expenses)
    ws.merge_cells(f"A{tr}:D{tr}")
    st(ws, tr, 1, f"TOTAL  ({len(expenses)} receipts)",
       bold=True, size=11, color="FFFFFF", bg="2E7D32", ha="center", bdr=B_TOP)
    st(ws, tr, 5, grand_total, bold=True, size=11, bg="E8F5E9", ha="right", fmt=PESO, bdr=B_TOP)
    for ci in [6, 7]:
        st(ws, tr, ci, "", bg="E8F5E9", bdr=B_TOP)
    ws.row_dimensions[tr].height = 20

    # 주석
    note_row = tr + 2
    ws.merge_cells(f"A{note_row}:G{note_row}")
    st(ws, note_row, 1,
       "⚠  Please verify all extracted amounts against original receipts before final submission.",
       size=8, color="AA5500", bg="FFFDE7", ha="left", bdr=None)

    ws.freeze_panes = "A5"

    # ══ Sheet 2: Category Summary ════════════════════════════════════════════
    ws2 = wb.create_sheet("Category Summary")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 16

    ws2.merge_cells("A1:D1")
    st(ws2, 1, 1, f"EXPENSE SUMMARY BY CATEGORY — {report_date}",
       bold=True, size=13, color="FFFFFF", bg="1B5E20", ha="center", bdr=B_MED)
    ws2.row_dimensions[1].height = 22

    for ci, h in enumerate(["Category", "Count", "Total (PHP)", "% of Total"], 1):
        st(ws2, 2, ci, h, bold=True, size=10, color="FFFFFF",
           bg="2E7D32", ha="center", bdr=B_MED)
    ws2.row_dimensions[2].height = 18

    # 카테고리별 집계
    cat_totals: dict = {}
    for exp in expenses:
        cat = exp.get("category", "other") or "other"
        if cat not in cat_totals:
            cat_totals[cat] = {"count": 0, "total": 0.0}
        cat_totals[cat]["count"] += 1
        cat_totals[cat]["total"] += exp.get("amount", 0) or 0

    sorted_cats = sorted(cat_totals.items(), key=lambda x: x[1]["total"], reverse=True)

    for ri, (cat, data) in enumerate(sorted_cats, 3):
        cat_label = CATEGORIES.get(cat, "📦 Other")
        cat_bg    = CATEGORY_COLORS.get(cat, "EEEEEE")
        pct = (data["total"] / grand_total * 100) if grand_total > 0 else 0

        st(ws2, ri, 1, cat_label,           size=10, ha="left",   bg=cat_bg)
        st(ws2, ri, 2, data["count"],        size=10, ha="center", bg=cat_bg)
        st(ws2, ri, 3, data["total"],        size=10, ha="right",  bg=cat_bg, fmt=PESO)
        st(ws2, ri, 4, f"{pct:.1f}%",       size=10, ha="center", bg=cat_bg)
        ws2.row_dimensions[ri].height = 16

    tr2 = 3 + len(sorted_cats)
    for ci, (val, fmt) in enumerate(
        [("GRAND TOTAL", None), (len(expenses), None), (grand_total, PESO), ("100%", None)], 1
    ):
        st(ws2, tr2, ci, val, bold=True, size=11, color="FFFFFF",
           bg="1B5E20", ha="right" if ci in (3,) else "center", fmt=fmt, bdr=B_MED)
    ws2.row_dimensions[tr2].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─── 엔드포인트 ──────────────────────────────────────────────────────────────

@router.post("/scan")
async def scan_receipts(files: List[UploadFile] = File(...)):
    """
    영수증 이미지를 Claude Vision으로 OCR하여 구조화된 JSON 반환.
    여러 장 동시 업로드 가능.
    """
    if not files:
        raise HTTPException(status_code=400, detail="파일을 선택해주세요")

    for f in files:
        if f.content_type not in ALLOWED_TYPES:
            raise HTTPException(
                status_code=400,
                detail=f"지원하지 않는 파일 형식: {f.filename} ({f.content_type})"
            )

    results = []
    errors  = []

    for i, upload in enumerate(files):
        try:
            content = await upload.read()
            logger.info(f"🧾 영수증 OCR: {upload.filename} ({len(content)//1024}KB)")
            data = await ocr_receipt(content, upload.content_type or "image/jpeg", upload.filename or "")
            data["id"]       = i + 1
            data["amount"]   = float(data.get("total_amount") or 0)
            data["category"] = data.get("category_guess") or "other"
            results.append(data)
            logger.info(
                f"✅ OCR 완료: {upload.filename} | "
                f"{data.get('vendor')} | PHP {data.get('amount', 0):.2f} | {data.get('category')}"
            )
        except json.JSONDecodeError as e:
            errors.append(f"{upload.filename}: JSON 파싱 실패")
            logger.error(f"JSON parse error {upload.filename}: {e}")
        except Exception as e:
            errors.append(f"{upload.filename}: {str(e)}")
            logger.error(f"OCR error {upload.filename}: {e}")

    if not results and errors:
        raise HTTPException(status_code=422, detail=f"OCR 실패: {'; '.join(errors)}")

    return {"results": results, "errors": errors, "count": len(results)}


@router.post("/export")
async def export_expense_excel(req: ExportRequest):
    """지출 데이터 → 포맷된 Excel 파일 다운로드"""
    if not req.expenses:
        raise HTTPException(status_code=400, detail="지출 데이터가 없습니다")

    expenses_dict = [e.model_dump() for e in req.expenses]
    excel_bytes   = build_expense_excel(req.report_date, expenses_dict)
    filename      = f"Expense_Report_{req.report_date}.xlsx"

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
