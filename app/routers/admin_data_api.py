"""
============================================================
📌 Admin Data Management API - FastAPI
📋 목적: 테라피스트, 예약, 드라이버 데이터 CRUD + Excel 내보내기
🔧 엔드포인트: /api/admin/data/*
📅 작성일: 2026-05-21
============================================================
"""

from fastapi import APIRouter, HTTPException, status
from fastapi.responses import FileResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os
import tempfile
from typing import List, Dict, Any

router = APIRouter(prefix="/api/admin/data", tags=["Admin Data"])

# ============================================================
# Mock 데이터
# ============================================================

THERAPISTS_DB = [
    {"id": 1, "name": "김수진", "speciality": "스웨디시", "experience": 5, "rating": 4.9, "status": "Active"},
    {"id": 2, "name": "이은정", "speciality": "타이마사지", "experience": 3, "rating": 4.8, "status": "Active"},
    {"id": 3, "name": "박민지", "speciality": "딥티슈", "experience": 7, "rating": 4.95, "status": "Active"},
    {"id": 4, "name": "최은미", "speciality": "핫스톤", "experience": 4, "rating": 4.7, "status": "Active"},
    {"id": 5, "name": "정유진", "speciality": "아로마테라피", "experience": 6, "rating": 4.85, "status": "Inactive"},
]

BOOKINGS_DB = [
    {"id": 1, "customer": "고객1", "therapist": "김수진", "service": "스웨디시", "date": "2026-05-21", "time": "14:00", "duration": 60, "status": "예약"},
    {"id": 2, "customer": "고객2", "therapist": "이은정", "service": "타이마사지", "date": "2026-05-21", "time": "15:30", "duration": 90, "status": "진행중"},
    {"id": 3, "customer": "고객3", "therapist": "박민지", "service": "딥티슈", "date": "2026-05-22", "time": "10:00", "duration": 60, "status": "완료"},
    {"id": 4, "customer": "고객4", "therapist": "최은미", "service": "핫스톤", "date": "2026-05-22", "time": "11:30", "duration": 75, "status": "예약"},
]

DRIVERS_DB = [
    {"id": 1, "name": "김철수", "vehicle": "검은색 쏘나타", "license": "12가1234", "phone": "010-1111-1111", "status": "운행중", "earnings": 285000},
    {"id": 2, "name": "이영수", "vehicle": "흰색 그랜저", "license": "12가5678", "phone": "010-2222-2222", "status": "대기중", "earnings": 195000},
    {"id": 3, "name": "박준호", "vehicle": "빨간색 K5", "license": "12가9999", "phone": "010-3333-3333", "status": "운행중", "earnings": 245000},
    {"id": 4, "name": "최준영", "vehicle": "은색 투산", "license": "12다1234", "phone": "010-4444-4444", "status": "대기중", "earnings": 160000},
]

# ============================================================
# Therapists API
# ============================================================

@router.get("/therapists")
async def get_therapists():
    """모든 테라피스트 조회"""
    return {
        "status": "success",
        "data": THERAPISTS_DB,
        "count": len(THERAPISTS_DB)
    }

@router.post("/therapists")
async def create_therapist(therapist: Dict[str, Any]):
    """새로운 테라피스트 추가"""
    new_id = max([t["id"] for t in THERAPISTS_DB]) + 1
    new_therapist = {
        "id": new_id,
        **therapist
    }
    THERAPISTS_DB.append(new_therapist)
    return {"status": "success", "data": new_therapist}

@router.put("/therapists/{therapist_id}")
async def update_therapist(therapist_id: int, therapist: Dict[str, Any]):
    """테라피스트 정보 수정"""
    for t in THERAPISTS_DB:
        if t["id"] == therapist_id:
            t.update(therapist)
            return {"status": "success", "data": t}
    raise HTTPException(status_code=404, detail="Therapist not found")

@router.delete("/therapists/{therapist_id}")
async def delete_therapist(therapist_id: int):
    """테라피스트 삭제"""
    global THERAPISTS_DB
    THERAPISTS_DB = [t for t in THERAPISTS_DB if t["id"] != therapist_id]
    return {"status": "success", "message": "Deleted"}

@router.post("/therapists/export")
async def export_therapists_excel():
    """테라피스트 정보를 Excel로 내보내기"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "테라피스트"

    # 헤더 설정
    headers = ["ID", "이름", "전문분야", "경력(년)", "평점", "상태"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # 데이터 입력
    for row, therapist in enumerate(THERAPISTS_DB, 2):
        ws.cell(row=row, column=1, value=therapist["id"])
        ws.cell(row=row, column=2, value=therapist["name"])
        ws.cell(row=row, column=3, value=therapist["speciality"])
        ws.cell(row=row, column=4, value=therapist["experience"])
        ws.cell(row=row, column=5, value=therapist["rating"])
        ws.cell(row=row, column=6, value=therapist["status"])

    # 열 너비 자동 조정
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 20)

    # 파일 저장
    filename = f"therapists_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(tempfile.gettempdir(), filename)
    wb.save(filepath)

    return FileResponse(filepath, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ============================================================
# Bookings API
# ============================================================

@router.get("/bookings")
async def get_bookings():
    """모든 예약 조회"""
    return {
        "status": "success",
        "data": BOOKINGS_DB,
        "count": len(BOOKINGS_DB)
    }

@router.post("/bookings")
async def create_booking(booking: Dict[str, Any]):
    """새로운 예약 추가"""
    new_id = max([b["id"] for b in BOOKINGS_DB]) + 1
    new_booking = {
        "id": new_id,
        **booking
    }
    BOOKINGS_DB.append(new_booking)
    return {"status": "success", "data": new_booking}

@router.put("/bookings/{booking_id}")
async def update_booking(booking_id: int, booking: Dict[str, Any]):
    """예약 정보 수정"""
    for b in BOOKINGS_DB:
        if b["id"] == booking_id:
            b.update(booking)
            return {"status": "success", "data": b}
    raise HTTPException(status_code=404, detail="Booking not found")

@router.delete("/bookings/{booking_id}")
async def delete_booking(booking_id: int):
    """예약 삭제"""
    global BOOKINGS_DB
    BOOKINGS_DB = [b for b in BOOKINGS_DB if b["id"] != booking_id]
    return {"status": "success", "message": "Deleted"}

@router.post("/bookings/export")
async def export_bookings_excel():
    """예약 정보를 Excel로 내보내기"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "예약"

    # 헤더 설정
    headers = ["ID", "고객", "테라피스트", "서비스", "날짜", "시간", "지속시간(분)", "상태"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # 데이터 입력
    for row, booking in enumerate(BOOKINGS_DB, 2):
        ws.cell(row=row, column=1, value=booking["id"])
        ws.cell(row=row, column=2, value=booking["customer"])
        ws.cell(row=row, column=3, value=booking["therapist"])
        ws.cell(row=row, column=4, value=booking["service"])
        ws.cell(row=row, column=5, value=booking["date"])
        ws.cell(row=row, column=6, value=booking["time"])
        ws.cell(row=row, column=7, value=booking["duration"])
        ws.cell(row=row, column=8, value=booking["status"])

    # 열 너비 자동 조정
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 20)

    # 파일 저장
    filename = f"bookings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(tempfile.gettempdir(), filename)
    wb.save(filepath)

    return FileResponse(filepath, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ============================================================
# Drivers API
# ============================================================

@router.get("/drivers")
async def get_drivers():
    """모든 드라이버 조회"""
    return {
        "status": "success",
        "data": DRIVERS_DB,
        "count": len(DRIVERS_DB)
    }

@router.post("/drivers")
async def create_driver(driver: Dict[str, Any]):
    """새로운 드라이버 추가"""
    new_id = max([d["id"] for d in DRIVERS_DB]) + 1
    new_driver = {
        "id": new_id,
        **driver
    }
    DRIVERS_DB.append(new_driver)
    return {"status": "success", "data": new_driver}

@router.put("/drivers/{driver_id}")
async def update_driver(driver_id: int, driver: Dict[str, Any]):
    """드라이버 정보 수정"""
    for d in DRIVERS_DB:
        if d["id"] == driver_id:
            d.update(driver)
            return {"status": "success", "data": d}
    raise HTTPException(status_code=404, detail="Driver not found")

@router.delete("/drivers/{driver_id}")
async def delete_driver(driver_id: int):
    """드라이버 삭제"""
    global DRIVERS_DB
    DRIVERS_DB = [d for d in DRIVERS_DB if d["id"] != driver_id]
    return {"status": "success", "message": "Deleted"}

@router.post("/drivers/export")
async def export_drivers_excel():
    """드라이버 정보를 Excel로 내보내기"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "드라이버"

    # 헤더 설정
    headers = ["ID", "이름", "차량", "면허번호", "휴대폰", "상태", "오늘수익(₩)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # 데이터 입력
    for row, driver in enumerate(DRIVERS_DB, 2):
        ws.cell(row=row, column=1, value=driver["id"])
        ws.cell(row=row, column=2, value=driver["name"])
        ws.cell(row=row, column=3, value=driver["vehicle"])
        ws.cell(row=row, column=4, value=driver["license"])
        ws.cell(row=row, column=5, value=driver["phone"])
        ws.cell(row=row, column=6, value=driver["status"])
        ws.cell(row=row, column=7, value=driver["earnings"])

    # 열 너비 자동 조정
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 20)

    # 파일 저장
    filename = f"drivers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(tempfile.gettempdir(), filename)
    wb.save(filepath)

    return FileResponse(filepath, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
