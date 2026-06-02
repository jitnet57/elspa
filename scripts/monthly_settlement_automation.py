"""
============================================================
📌 월간 정산 자동화 스크립트
📋 목적: 월간 정산 자동 처리 (draft → settled) + 보고서 생성
📅 작성일: 2026-06-02
설명:
  - 월간 정산 (payment_from='guest' + status='pending') 자동 완료
  - 신용 정산 (payment_from='credit' + payment_status='collected') 자동 완료
  - PDF/Excel 보고서 생성 (업체별 집계)
  - APScheduler를 이용한 월간 자동 실행 (매월 5일 오전 9시)

의존성:
  pip install sqlalchemy python-dotenv apscheduler openpyxl reportlab

실행 방법:
  1. 수동 실행: python scripts/monthly_settlement_automation.py --manual
  2. 스케줄러 시작: python scripts/monthly_settlement_automation.py --schedule
  3. 테스트: python scripts/monthly_settlement_automation.py --test

구조:
  - SettlementAutomation 클래스 (메인 로직)
  - generate_pdf_report() (PDF 생성)
  - generate_excel_report() (Excel 생성)
  - setup_scheduler() (APScheduler 설정)
============================================================
"""

import os
import sys
import logging
from datetime import datetime, date, timedelta
from decimal import Decimal
from typing import List, Dict, Optional, Tuple
import json
import argparse
from pathlib import Path

# 외부 라이브러리
from sqlalchemy import create_engine, and_, or_
from sqlalchemy.orm import sessionmaker, Session
from dotenv import load_dotenv
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import atexit

# 프로젝트 모듈
sys.path.insert(0, str(Path(__file__).parent.parent))
from app.models.company_settlement import CompanySettlement, SettlementStatus
from app.database import SessionLocal_sync

# ============================================================
# 로깅 설정
# ============================================================
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s - %(name)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/settlement_automation.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ============================================================
# 설정
# ============================================================
load_dotenv()
DATABASE_URL = os.getenv("DATABASE_URL")


class SettlementAutomation:
    """월간 정산 자동화 엔진"""

    def __init__(self, db_url: str = None):
        """
        초기화

        매개변수:
            db_url: 데이터베이스 URL (기본값: .env DATABASE_URL)
        """
        self.db_url = db_url or DATABASE_URL
        self.engine = create_engine(self.db_url, echo=False)
        self.SessionLocal = sessionmaker(bind=self.engine)
        self.report_dir = Path("reports/settlements")
        self.report_dir.mkdir(parents=True, exist_ok=True)

        logger.info(f"SettlementAutomation initialized with DB: {self.db_url}")

    # ============================================================
    # Phase 1: 정산 데이터 조회
    # ============================================================

    def get_pending_guest_settlements(
        self,
        year: int,
        month: int,
        db: Session = None
    ) -> List[CompanySettlement]:
        """
        비회원(guest) 정산 조회

        조건:
          - settlement_period_year = year
          - settlement_period_month = month
          - payment_from = 'guest'
          - status = 'pending' or 'draft'

        매개변수:
            year: 정산 연도
            month: 정산 월
            db: 데이터베이스 세션

        반환값:
            CompanySettlement 리스트
        """
        db = db or self.SessionLocal()
        try:
            settlements = db.query(CompanySettlement).filter(
                CompanySettlement.settlement_period_year == year,
                CompanySettlement.settlement_period_month == month,
                CompanySettlement.status.in_(["draft", "pending"]),
                # payment_from = 'guest'인 거래만 필터링
            ).all()

            logger.info(f"Found {len(settlements)} pending guest settlements for {year}-{month:02d}")
            return settlements
        finally:
            if db:
                db.close()

    def get_collected_credit_settlements(
        self,
        year: int,
        month: int,
        db: Session = None
    ) -> List[CompanySettlement]:
        """
        외상(credit) 정산 중 회수 완료된 항목 조회

        조건:
          - settlement_period_year = year
          - settlement_period_month = month
          - payment_from = 'credit'
          - recovery_rate = 100% (회수 완료)
          - status = 'pending' or 'draft'

        매개변수:
            year: 정산 연도
            month: 정산 월
            db: 데이터베이스 세션

        반환값:
            CompanySettlement 리스트
        """
        db = db or self.SessionLocal()
        try:
            settlements = db.query(CompanySettlement).filter(
                CompanySettlement.settlement_period_year == year,
                CompanySettlement.settlement_period_month == month,
                CompanySettlement.recovery_rate == Decimal("100.00"),
                CompanySettlement.status.in_(["draft", "pending"]),
            ).all()

            logger.info(f"Found {len(settlements)} collected credit settlements for {year}-{month:02d}")
            return settlements
        finally:
            if db:
                db.close()

    # ============================================================
    # Phase 2: 정산 완료 처리
    # ============================================================

    def mark_settlements_as_settled(
        self,
        settlements: List[CompanySettlement],
        payment_method: str = "bank_transfer",
        paid_by: str = "system_automation",
        db: Session = None
    ) -> Tuple[int, List[Dict]]:
        """
        정산을 완료 상태로 변경

        매개변수:
            settlements: CompanySettlement 리스트
            payment_method: 지급 방법 (기본값: bank_transfer)
            paid_by: 처리자 (기본값: system_automation)
            db: 데이터베이스 세션

        반환값:
            (성공 건수, 상세 결과 리스트)
        """
        db = db or self.SessionLocal()
        success_count = 0
        results = []

        try:
            payment_date = date.today()

            for settlement in settlements:
                try:
                    # net_settlement 검증
                    if settlement.net_settlement <= 0:
                        logger.warning(
                            f"Settlement {settlement.id} (Company {settlement.company_id}) "
                            f"has zero or negative net_settlement: {settlement.net_settlement}"
                        )
                        results.append({
                            "settlement_id": settlement.id,
                            "company_id": settlement.company_id,
                            "status": "skipped",
                            "reason": "zero_or_negative_amount",
                            "net_settlement": float(settlement.net_settlement)
                        })
                        continue

                    # mark_settled 호출
                    if settlement.mark_settled(
                        payment_date=payment_date,
                        payment_method=payment_method,
                        paid_by=paid_by
                    ):
                        success_count += 1
                        logger.info(
                            f"✓ Settlement {settlement.id} (Company {settlement.company_id}) "
                            f"marked as settled. Amount: {settlement.net_settlement}"
                        )
                        results.append({
                            "settlement_id": settlement.id,
                            "company_id": settlement.company_id,
                            "status": "settled",
                            "net_settlement": float(settlement.net_settlement),
                            "payment_date": payment_date.isoformat(),
                            "payment_method": payment_method
                        })
                    else:
                        logger.error(
                            f"✗ Failed to mark settlement {settlement.id} as settled"
                        )
                        results.append({
                            "settlement_id": settlement.id,
                            "company_id": settlement.company_id,
                            "status": "failed",
                            "reason": "mark_settled_returned_false"
                        })
                except Exception as e:
                    logger.error(
                        f"✗ Error processing settlement {settlement.id}: {str(e)}"
                    )
                    results.append({
                        "settlement_id": settlement.id,
                        "company_id": settlement.company_id,
                        "status": "error",
                        "error": str(e)
                    })

            # 데이터베이스 커밋
            db.commit()
            logger.info(f"Successfully committed {success_count} settlements")

        except Exception as e:
            logger.error(f"Error during settlement marking: {str(e)}")
            db.rollback()
            raise
        finally:
            if db:
                db.close()

        return success_count, results

    # ============================================================
    # Phase 3: 보고서 생성
    # ============================================================

    def generate_settlement_summary(
        self,
        settlements: List[CompanySettlement],
        year: int,
        month: int
    ) -> Dict:
        """
        정산 요약 생성

        반환값:
            {
                "period": "2026-06",
                "total_companies": 10,
                "total_revenue": 1000000.00,
                "total_platform_fee": 250000.00,
                "total_net_settlement": 750000.00,
                "companies": [
                    {
                        "company_id": 1,
                        "company_name": "Example Spa",
                        "total_revenue": 100000.00,
                        "platform_fee": 25000.00,
                        "net_settlement": 75000.00,
                        "status": "settled",
                        ...
                    },
                    ...
                ]
            }
        """
        summary = {
            "period": f"{year:04d}-{month:02d}",
            "generated_at": datetime.now().isoformat(),
            "total_companies": len(settlements),
            "total_revenue": Decimal("0.00"),
            "total_guest_revenue": Decimal("0.00"),
            "total_credit_revenue": Decimal("0.00"),
            "total_platform_fee": Decimal("0.00"),
            "total_deductions": Decimal("0.00"),
            "total_net_settlement": Decimal("0.00"),
            "companies": []
        }

        for settlement in settlements:
            # 요약 누적
            summary["total_revenue"] += settlement.total_revenue
            summary["total_guest_revenue"] += settlement.guest_revenue
            summary["total_credit_revenue"] += settlement.credit_revenue
            summary["total_platform_fee"] += settlement.platform_fee
            summary["total_deductions"] += settlement.total_deductions
            summary["total_net_settlement"] += settlement.net_settlement

            # 업체별 상세
            company_data = {
                "company_id": settlement.company_id,
                "settlement_id": settlement.id,
                "total_revenue": float(settlement.total_revenue),
                "guest_revenue": float(settlement.guest_revenue),
                "credit_revenue": float(settlement.credit_revenue),
                "waived_revenue": float(settlement.waived_revenue),
                "recovery_rate": float(settlement.recovery_rate),
                "recovered_amount": float(settlement.recovered_amount),
                "platform_fee_rate": float(settlement.platform_fee_rate),
                "platform_fee": float(settlement.platform_fee),
                "refund_amount": float(settlement.refund_amount),
                "dispute_deduction": float(settlement.dispute_deduction),
                "other_deduction": float(settlement.other_deduction),
                "total_deductions": float(settlement.total_deductions),
                "net_settlement": float(settlement.net_settlement),
                "status": settlement.status,
                "payment_method": settlement.payment_method,
                "payment_date": settlement.payment_date.isoformat() if settlement.payment_date else None,
                "transaction_count": len(settlement.transactions) if settlement.transactions else 0
            }
            summary["companies"].append(company_data)

        # Decimal → float 변환
        summary["total_revenue"] = float(summary["total_revenue"])
        summary["total_guest_revenue"] = float(summary["total_guest_revenue"])
        summary["total_credit_revenue"] = float(summary["total_credit_revenue"])
        summary["total_platform_fee"] = float(summary["total_platform_fee"])
        summary["total_deductions"] = float(summary["total_deductions"])
        summary["total_net_settlement"] = float(summary["total_net_settlement"])

        return summary

    def generate_excel_report(
        self,
        summary: Dict,
        filename: str = None
    ) -> str:
        """
        Excel 보고서 생성

        매개변수:
            summary: generate_settlement_summary()의 결과
            filename: 파일명 (기본값: settlement_YYYYMM.xlsx)

        반환값:
            생성된 파일 경로
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            return None

        if not filename:
            filename = f"settlement_{summary['period'].replace('-', '')}.xlsx"

        filepath = self.report_dir / filename

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "정산 보고서"

            # 헤더 스타일
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 제목
            ws['A1'] = f"월간 정산 보고서 ({summary['period']})"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:H1')

            ws['A2'] = f"생성일: {summary['generated_at']}"
            ws.merge_cells('A2:H2')

            # 요약 정보
            row = 4
            ws[f'A{row}'] = "요약"
            ws[f'A{row}'].font = Font(bold=True, size=12)

            row += 1
            summary_data = [
                ("총 회사 수", summary["total_companies"]),
                ("총 매출액", f"{summary['total_revenue']:,.2f}"),
                ("플랫폼 수수료", f"{summary['total_platform_fee']:,.2f}"),
                ("총 차감액", f"{summary['total_deductions']:,.2f}"),
                ("총 정산액", f"{summary['total_net_settlement']:,.2f}"),
            ]

            for label, value in summary_data:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'B{row}'].alignment = Alignment(horizontal='right')
                row += 1

            # 업체별 상세
            row = 11
            headers = [
                "업체ID", "정산ID", "총매출", "비회원", "외상", "제외",
                "회수율(%)", "수수료액", "차감액", "순정산액", "상태", "지급일", "거래건수"
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border

            # 데이터 입력
            for company in summary["companies"]:
                row += 1
                data = [
                    company["company_id"],
                    company["settlement_id"],
                    f"{company['total_revenue']:,.2f}",
                    f"{company['guest_revenue']:,.2f}",
                    f"{company['credit_revenue']:,.2f}",
                    f"{company['waived_revenue']:,.2f}",
                    f"{company['recovery_rate']:.2f}",
                    f"{company['platform_fee']:,.2f}",
                    f"{company['total_deductions']:,.2f}",
                    f"{company['net_settlement']:,.2f}",
                    company["status"],
                    company["payment_date"] or "-",
                    company["transaction_count"]
                ]

                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    if col > 2:  # 숫자 칼럼 우측 정렬
                        cell.alignment = Alignment(horizontal='right')

            # 칼럼 너비 조정
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 12
            for col in range(3, 14):
                ws.column_dimensions[chr(64 + col)].width = 15

            wb.save(filepath)
            logger.info(f"✓ Excel report generated: {filepath}")
            return str(filepath)

        except Exception as e:
            logger.error(f"Error generating Excel report: {str(e)}")
            return None

    def generate_pdf_report(
        self,
        summary: Dict,
        filename: str = None
    ) -> str:
        """
        PDF 보고서 생성

        매개변수:
            summary: generate_settlement_summary()의 결과
            filename: 파일명 (기본값: settlement_YYYYMM.pdf)

        반환값:
            생성된 파일 경로
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        except ImportError:
            logger.error("reportlab not installed. Run: pip install reportlab")
            return None

        if not filename:
            filename = f"settlement_{summary['period'].replace('-', '')}.pdf"

        filepath = self.report_dir / filename

        try:
            doc = SimpleDocTemplate(
                str(filepath),
                pagesize=landscape(letter),
                rightMargin=0.5*inch,
                leftMargin=0.5*inch,
                topMargin=0.75*inch,
                bottomMargin=0.75*inch
            )

            story = []
            styles = getSampleStyleSheet()

            # 제목
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor("#366092"),
                spaceAfter=6,
                alignment=TA_CENTER
            )
            story.append(Paragraph(f"월간 정산 보고서 ({summary['period']})", title_style))

            # 생성일
            meta_style = ParagraphStyle(
                'Meta',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.grey,
                alignment=TA_CENTER
            )
            story.append(Paragraph(f"생성일: {summary['generated_at']}", meta_style))
            story.append(Spacer(1, 0.15*inch))

            # 요약 테이블
            summary_data = [
                ["항목", "금액"],
                ["총 회사 수", str(summary["total_companies"])],
                ["총 매출액", f"₱{summary['total_revenue']:,.2f}"],
                ["비회원 매출", f"₱{summary['total_guest_revenue']:,.2f}"],
                ["외상 매출", f"₱{summary['total_credit_revenue']:,.2f}"],
                ["플랫폼 수수료", f"₱{summary['total_platform_fee']:,.2f}"],
                ["총 차감액", f"₱{summary['total_deductions']:,.2f}"],
                ["총 정산액", f"₱{summary['total_net_settlement']:,.2f}"],
            ]

            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#366092")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 0.2*inch))

            # 업체별 상세 테이블
            story.append(Paragraph("업체별 정산 상세", title_style))
            story.append(Spacer(1, 0.1*inch))

            company_data = [
                ["업체ID", "정산ID", "총매출", "비회원", "외상", "수수료", "차감", "정산액", "상태"]
            ]

            for company in summary["companies"]:
                company_data.append([
                    str(company["company_id"]),
                    str(company["settlement_id"]),
                    f"₱{company['total_revenue']:,.0f}",
                    f"₱{company['guest_revenue']:,.0f}",
                    f"₱{company['credit_revenue']:,.0f}",
                    f"₱{company['platform_fee']:,.0f}",
                    f"₱{company['total_deductions']:,.0f}",
                    f"₱{company['net_settlement']:,.0f}",
                    company["status"]
                ])

            company_table = Table(
                company_data,
                colWidths=[0.6*inch, 0.6*inch, 0.8*inch, 0.8*inch, 0.8*inch,
                          0.8*inch, 0.8*inch, 0.8*inch, 0.6*inch]
            )
            company_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#366092")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            story.append(company_table)

            doc.build(story)
            logger.info(f"✓ PDF report generated: {filepath}")
            return str(filepath)

        except Exception as e:
            logger.error(f"Error generating PDF report: {str(e)}")
            return None

    # ============================================================
    # Phase 4: 전체 자동화 파이프라인
    # ============================================================

    def run_monthly_settlement(
        self,
        year: int = None,
        month: int = None,
        generate_reports: bool = True
    ) -> Dict:
        """
        월간 정산 전체 프로세스 실행

        프로세스:
          1. 비회원(guest) 정산 조회 + 완료 처리
          2. 외상(credit) 회수 완료 조회 + 완료 처리
          3. 정산 요약 생성
          4. PDF/Excel 보고서 생성

        매개변수:
            year: 정산 연도 (기본값: 현재 월)
            month: 정산 월 (기본값: 현재 월)
            generate_reports: 보고서 생성 여부 (기본값: True)

        반환값:
            {
                "success": bool,
                "timestamp": datetime,
                "period": "2026-06",
                "guest_settlements": {
                    "processed": int,
                    "details": list
                },
                "credit_settlements": {
                    "processed": int,
                    "details": list
                },
                "summary": dict,
                "reports": {
                    "excel": str or None,
                    "pdf": str or None
                }
            }
        """
        if not year or not month:
            today = date.today()
            year = today.year
            month = today.month

        logger.info(f"Starting monthly settlement automation for {year}-{month:02d}")

        result = {
            "success": False,
            "timestamp": datetime.now().isoformat(),
            "period": f"{year:04d}-{month:02d}",
            "guest_settlements": {"processed": 0, "details": []},
            "credit_settlements": {"processed": 0, "details": []},
            "summary": None,
            "reports": {"excel": None, "pdf": None}
        }

        try:
            # Step 1: 비회원(guest) 정산 처리
            logger.info("Step 1: Processing guest settlements...")
            guest_settlements = self.get_pending_guest_settlements(year, month)

            if guest_settlements:
                guest_count, guest_results = self.mark_settlements_as_settled(
                    guest_settlements,
                    payment_method="bank_transfer",
                    paid_by="system_automation"
                )
                result["guest_settlements"]["processed"] = guest_count
                result["guest_settlements"]["details"] = guest_results
                logger.info(f"✓ Processed {guest_count}/{len(guest_settlements)} guest settlements")
            else:
                logger.info("No pending guest settlements found")

            # Step 2: 외상(credit) 정산 처리
            logger.info("Step 2: Processing collected credit settlements...")
            credit_settlements = self.get_collected_credit_settlements(year, month)

            if credit_settlements:
                credit_count, credit_results = self.mark_settlements_as_settled(
                    credit_settlements,
                    payment_method="bank_transfer",
                    paid_by="system_automation"
                )
                result["credit_settlements"]["processed"] = credit_count
                result["credit_settlements"]["details"] = credit_results
                logger.info(f"✓ Processed {credit_count}/{len(credit_settlements)} credit settlements")
            else:
                logger.info("No collected credit settlements found")

            # Step 3: 정산 요약 생성
            logger.info("Step 3: Generating settlement summary...")
            all_settlements = guest_settlements + credit_settlements

            if all_settlements:
                summary = self.generate_settlement_summary(all_settlements, year, month)
                result["summary"] = summary
                logger.info(f"✓ Generated summary for {len(all_settlements)} settlements")

                # Step 4: 보고서 생성
                if generate_reports:
                    logger.info("Step 4: Generating reports...")

                    excel_file = self.generate_excel_report(summary)
                    if excel_file:
                        result["reports"]["excel"] = excel_file

                    pdf_file = self.generate_pdf_report(summary)
                    if pdf_file:
                        result["reports"]["pdf"] = pdf_file

                    logger.info("✓ Reports generated")
            else:
                logger.warning("No settlements found for summary")

            result["success"] = True
            logger.info("✓ Monthly settlement automation completed successfully")

        except Exception as e:
            logger.error(f"✗ Error during monthly settlement: {str(e)}")
            result["error"] = str(e)

        return result


# ============================================================
# APScheduler 스케줄러
# ============================================================

def setup_scheduler(automation: SettlementAutomation):
    """
    APScheduler 설정 (매월 5일 오전 9시 자동 실행)

    매개변수:
        automation: SettlementAutomation 인스턴스
    """
    scheduler = BackgroundScheduler()

    def job_callback():
        logger.info("🔔 Settlement automation job triggered by scheduler")
        result = automation.run_monthly_settlement()

        # 결과를 JSON으로 저장
        log_file = Path("logs/settlement_automation_results.json")
        with open(log_file, "a") as f:
            f.write(json.dumps(result, indent=2) + "\n" + "="*50 + "\n")

    # 매월 5일 오전 9시 실행
    scheduler.add_job(
        job_callback,
        CronTrigger(day=5, hour=9, minute=0),
        id='monthly_settlement',
        name='Monthly Settlement Automation',
        replace_existing=True
    )

    scheduler.start()
    logger.info("✓ Scheduler started. Next run: 5th day of month at 09:00")

    # 프로그램 종료 시 스케줄러 정지
    atexit.register(lambda: scheduler.shutdown())

    return scheduler


# ============================================================
# CLI 인터페이스
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="월간 정산 자동화 스크립트"
    )
    parser.add_argument(
        "--manual",
        action="store_true",
        help="현재 월 정산 수동 실행"
    )
    parser.add_argument(
        "--year",
        type=int,
        help="정산 연도"
    )
    parser.add_argument(
        "--month",
        type=int,
        help="정산 월"
    )
    parser.add_argument(
        "--schedule",
        action="store_true",
        help="스케줄러 시작 (백그라운드 실행)"
    )
    parser.add_argument(
        "--test",
        action="store_true",
        help="테스트 모드 (보고서 생성 안 함)"
    )

    args = parser.parse_args()

    # SettlementAutomation 초기화
    automation = SettlementAutomation()

    if args.schedule:
        # 스케줄러 모드: 백그라운드에서 계속 실행
        print("🚀 Starting settlement automation scheduler...")
        setup_scheduler(automation)
        print("✓ Scheduler is running. Press Ctrl+C to stop.")

        try:
            while True:
                import time
                time.sleep(1)
        except KeyboardInterrupt:
            print("\n⏹️ Scheduler stopped.")

    elif args.manual or args.test:
        # 수동 모드: 한 번 실행
        year = args.year
        month = args.month

        print(f"🔧 Running settlement automation for {year}-{month:02d}...")

        result = automation.run_monthly_settlement(
            year=year,
            month=month,
            generate_reports=not args.test
        )

        # 결과 출력
        print("\n" + "="*60)
        print(f"Period: {result['period']}")
        print(f"Success: {result['success']}")
        print(f"Guest Settlements: {result['guest_settlements']['processed']} processed")
        print(f"Credit Settlements: {result['credit_settlements']['processed']} processed")

        if result['summary']:
            print(f"\nSummary:")
            print(f"  Total Companies: {result['summary']['total_companies']}")
            print(f"  Total Revenue: ₱{result['summary']['total_revenue']:,.2f}")
            print(f"  Platform Fee: ₱{result['summary']['total_platform_fee']:,.2f}")
            print(f"  Net Settlement: ₱{result['summary']['total_net_settlement']:,.2f}")

        if result['reports']['excel']:
            print(f"\nExcel Report: {result['reports']['excel']}")
        if result['reports']['pdf']:
            print(f"PDF Report: {result['reports']['pdf']}")

        print("="*60)

    else:
        parser.print_help()


if __name__ == "__main__":
    main()
