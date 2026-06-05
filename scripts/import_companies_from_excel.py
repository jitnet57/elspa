#!/usr/bin/env python3
"""
업체(Company) 데이터를 Excel에서 추출해서 Supabase에 등록
"""

import os
import sys
from dotenv import load_dotenv
from supabase import create_client
import openpyxl
from pathlib import Path

load_dotenv()
supabase = create_client(os.getenv("SUPABASE_URL"), os.getenv("SUPABASE_KEY"))

# Excel 파일 경로
excel_files = [
    Path.home() / "Downloads" / "02월 siw 마사지..xlsx",
    Path.home() / "Downloads" / "02월 가이드맨..xlsx",
    Path.home() / "Downloads" / "02월 투어스타..xlsx",
    Path.home() / "Downloads" / "WEGO com.xlsx",
    Path.home() / "Downloads" / "JTB.xlsx",
    Path.home() / "Downloads" / "HIS.xlsx",
]

companies_set = set()  # 중복 제거용

for excel_file in excel_files:
    if not excel_file.exists():
        print(f"⏭️  {excel_file.name} (파일 없음)")
        continue
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        print(f"\n📄 {excel_file.name}")
        
        # 첫 행부터 시작 (헤더 찾기)
        header_row = None
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            # "Company", "업체", "투어사", "여행사" 등의 컬럼 찾기
            if any(col and ("company" in str(col).lower() or "업체" in str(col) or "여행사" in str(col)) for col in row):
                header_row = row_idx
                break
        
        if not header_row:
            # 첫 번째 데이터 행 찾기
            header_row = 1
        
        # 데이터 추출
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx <= header_row:
                continue
            
            # 첫 번째 컬럼이 업체명일 가능성 높음
            company_name = row[0]
            
            if company_name and isinstance(company_name, str) and len(company_name.strip()) > 0:
                company_name = company_name.strip()
                if company_name not in companies_set and len(company_name) > 2:
                    companies_set.add(company_name)
                    print(f"  - {company_name}")
    
    except Exception as e:
        print(f"❌ {excel_file.name}: {str(e)[:50]}")

# Supabase에 등록
print(f"\n🔄 {len(companies_set)}개 업체 등록 시작...")

for company_name in sorted(companies_set):
    try:
        # 중복 확인
        existing = supabase.table("companies").select("id").eq("name", company_name).execute()
        if existing.data:
            print(f"⏭️  {company_name} (이미 있음)")
            continue
        
        # 새 업체 추가
        supabase.table("companies").insert({"name": company_name}).execute()
        print(f"✅ {company_name}")
    except Exception as e:
        print(f"⚠️  {company_name}: {str(e)[:40]}")

print("\n✅ 업체 등록 완료!")
