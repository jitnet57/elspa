"""
============================================================
📌 Sample Excel Files Generator
📋 목적: 테스트용 샘플 엑셀 파일 생성
🔧 파일:
   - bookings.xlsx (유효 + 에러 포함)
   - employees.xlsx (중복 데이터 포함)
   - expenses.xlsx (혼합 데이터 타입)
   - therapists.xlsx (완전 정상 데이터)
📅 작성일: 2026-06-02
============================================================
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, date, timedelta
from pathlib import Path

# 출력 디렉토리
OUTPUT_DIR = Path(__file__).parent / "sample_excel_files"
OUTPUT_DIR.mkdir(exist_ok=True)


# ============================================================
# 📌 스타일 유틸리티
# ============================================================

def get_header_style():
    """헤더 스타일"""
    return {
        'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
        'font': Font(bold=True, color='FFFFFF'),
        'alignment': Alignment(horizontal='center', vertical='center')
    }


def get_error_style():
    """에러 행 스타일"""
    return {
        'fill': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        'font': Font(color='9C0006')
    }


def apply_style(cell, style):
    """셀에 스타일 적용"""
    for key, value in style.items():
        if key == 'fill':
            cell.fill = value
        elif key == 'font':
            cell.font = value
        elif key == 'alignment':
            cell.alignment = value


# ============================================================
# 📌 1. Bookings.xlsx (예약 데이터 - 유효 + 에러)
# ============================================================

def create_bookings_excel():
    """예약 데이터 엑셀 (유효 + 에러 포함)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bookings"

    # 헤더
    headers = [
        "Booking ID",
        "Customer Name",
        "Therapist Name",
        "Service Type",
        "Booking Date",
        "Booking Time",
        "Duration (Minutes)",
        "Price",
        "Status"
    ]

    ws.append(headers)

    # 헤더 스타일 적용
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        apply_style(cell, get_header_style())

    # 데이터 (유효)
    valid_data = [
        [1, "김재준", "이순신", "Full Body Massage", datetime(2025, 6, 1), "10:00", 60, 80000, "Completed"],
        [2, "박지현", "강감찬", "Foot Massage", datetime(2025, 6, 2), "14:30", 30, 40000, "Completed"],
        [3, "최민우", "세종대왕", "Full Body Massage", datetime(2025, 6, 3), "15:00", 90, 120000, "Completed"],
        [4, "이순희", "이순신", "Shoulder Massage", datetime(2025, 6, 4), "11:00", 45, 60000, "Pending"],
        [5, "정수진", "강감찬", "Full Body Massage", datetime(2025, 6, 5), "13:00", 60, 80000, "Confirmed"],
    ]

    for idx, row in enumerate(valid_data, start=2):
        ws.append(row)

    # 에러 데이터 (주석: 각 에러 설명)
    error_data = [
        # Row 8: 예약 ID 없음 -> 필수 필드 에러
        [None, "임정민", "연개소문", "Full Body Massage", datetime(2025, 6, 6), "16:00", 60, 100000, "Pending"],

        # Row 9: 고객명 없음 -> 필수 필드 에러
        [6, None, "을지문덕", "Foot Massage", datetime(2025, 6, 7), "09:00", 30, 40000, "Pending"],

        # Row 10: 가격이 음수 -> 데이터 타입/범위 에러
        [7, "류성룡", "광개토대왕", "Full Body Massage", datetime(2025, 6, 8), "17:00", 60, -50000, "Pending"],

        # Row 11: 시간 형식 잘못됨 -> 포맷 에러
        [8, "신사임당", "이순신", "Full Body Massage", datetime(2025, 6, 9), "25:00", 60, 80000, "Pending"],

        # Row 12: 기간이 유효하지 않은 값 -> 범위 에러
        [9, "정약용", "강감찬", "Full Body Massage", datetime(2025, 6, 10), "18:00", 0, 80000, "Pending"],
    ]

    for idx, row in enumerate(error_data, start=8):
        ws.append(row)
        # 에러 행 스타일 적용
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=idx, column=col_num)
            apply_style(cell, get_error_style())

    # 열 너비 조정
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 12

    # 저장
    output_path = OUTPUT_DIR / "bookings.xlsx"
    wb.save(output_path)
    print(f"✅ 생성: {output_path}")


# ============================================================
# 📌 2. Employees.xlsx (직원 데이터 - 중복 포함)
# ============================================================

def create_employees_excel():
    """직원 데이터 엑셀 (중복 포함)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"

    # 헤더
    headers = ["Employee ID", "Name", "Phone", "Employee Type", "Pay Group", "Base Salary", "Hire Date"]
    ws.append(headers)

    # 헤더 스타일
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        apply_style(cell, get_header_style())

    # 데이터
    data = [
        # 유효한 데이터
        [1, "김철수", "010-1234-5678", "therapist", "weekly", 2500000, date(2024, 1, 15)],
        [2, "이영미", "010-2345-6789", "nail", "biweekly", 1800000, date(2024, 2, 20)],
        [3, "박준호", "010-3456-7890", "driver", "weekly", 2000000, date(2024, 3, 10)],

        # 중복 데이터 (같은 전화번호)
        [4, "김철수", "010-1234-5678", "therapist", "weekly", 2500000, date(2024, 1, 15)],  # 중복
        [5, "이영미", "010-2345-6789", "nail", "biweekly", 1800000, date(2024, 2, 20)],  # 중복

        # 더 많은 유효한 데이터
        [6, "최민지", "010-4567-8901", "therapist", "biweekly", 2700000, date(2024, 4, 25)],
        [7, "정은희", "010-5678-9012", "therapist", "weekly", 2500000, date(2024, 5, 5)],

        # 중복 (ID만 다름)
        [8, "박준호", "010-3456-7890", "driver", "weekly", 2000000, date(2024, 3, 10)],  # 중복
    ]

    for idx, row in enumerate(data, start=2):
        ws.append(row)
        # 중복 행 표시
        if idx in [5, 6, 9]:  # 중복 행 번호
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=idx, column=col_num)
                # 경고 스타일 (황색)
                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

    # 열 너비
    for col, width in enumerate(['A', 'B', 'C', 'D', 'E', 'F', 'G'], 1):
        ws.column_dimensions[width].width = 15

    output_path = OUTPUT_DIR / "employees.xlsx"
    wb.save(output_path)
    print(f"✅ 생성: {output_path}")


# ============================================================
# 📌 3. Expenses.xlsx (지출 데이터 - 혼합 타입)
# ============================================================

def create_expenses_excel():
    """지출 데이터 엑셀 (혼합 데이터 타입)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    # 헤더
    headers = ["Expense ID", "Category", "Description", "Amount", "Tax Rate", "Is Deductible", "Expense Date"]
    ws.append(headers)

    # 헤더 스타일
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        apply_style(cell, get_header_style())

    # 혼합 타입 데이터
    data = [
        # 모두 정상적인 타입
        [1, "Supplies", "Office supplies", 150000, 0.10, True, date(2025, 6, 1)],

        # 문자열로 된 숫자
        [2, "Utilities", "전기 요금", "250000", "0.10", "Yes", "2025-06-02"],

        # 혼합 boolean 표현
        [3, "Maintenance", "정기 점검", 180000, 0.05, 1, date(2025, 6, 3)],

        # 다양한 날짜 형식
        [4, "Training", "직원 교육", 500000, 0.15, "Y", datetime(2025, 6, 4, 10, 30)],

        # 소수점 수량
        [5, "Equipment", "장비 구매", 1500000.50, 0.20, False, date(2025, 6, 5)],

        # 과학 표기법 (매우 큰 수)
        [6, "Rent", "사무실 임차료", 3000000, 0.0, "no", date(2025, 6, 6)],

        # 음수 값 (환불/차용금)
        [7, "Refund", "고객 환불", -100000, 0, True, date(2025, 6, 7)],

        # 정상 데이터
        [8, "Marketing", "광고비", 250000, 0.10, True, date(2025, 6, 8)],
    ]

    for idx, row in enumerate(data, start=2):
        ws.append(row)

    # 열 너비
    for col, width in [('A', 12), ('B', 15), ('C', 20), ('D', 12), ('E', 10), ('F', 15), ('G', 12)]:
        ws.column_dimensions[col].width = width

    output_path = OUTPUT_DIR / "expenses.xlsx"
    wb.save(output_path)
    print(f"✅ 생성: {output_path}")


# ============================================================
# 📌 4. Therapists.xlsx (완전 정상 데이터)
# ============================================================

def create_therapists_excel():
    """테라피스트 데이터 엑셀 (모두 정상)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Therapists"

    # 헤더
    headers = ["Therapist ID", "Name", "Phone", "Specialization", "Hourly Rate", "License Number", "Hire Date"]
    ws.append(headers)

    # 헤더 스타일
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        apply_style(cell, get_header_style())

    # 모두 정상적인 데이터
    data = [
        [1, "이순신", "010-1111-1111", "Full Body Massage", 50000, "LIC-001", date(2023, 1, 10)],
        [2, "강감찬", "010-2222-2222", "Foot Massage", 40000, "LIC-002", date(2023, 2, 15)],
        [3, "세종대왕", "010-3333-3333", "Aromatherapy", 45000, "LIC-003", date(2023, 3, 20)],
        [4, "을지문덕", "010-4444-4444", "Sports Massage", 55000, "LIC-004", date(2023, 4, 25)],
        [5, "광개토대왕", "010-5555-5555", "Therapeutic Massage", 60000, "LIC-005", date(2023, 5, 30)],
        [6, "연개소문", "010-6666-6666", "Hot Stone Massage", 50000, "LIC-006", date(2023, 6, 5)],
        [7, "신사임당", "010-7777-7777", "Reflexology", 45000, "LIC-007", date(2023, 7, 10)],
        [8, "정약용", "010-8888-8888", "Swedish Massage", 50000, "LIC-008", date(2023, 8, 15)],
        [9, "류성룡", "010-9999-9999", "Thai Massage", 55000, "LIC-009", date(2023, 9, 20)],
        [10, "임정민", "010-1010-1010", "Full Body Massage", 50000, "LIC-010", date(2023, 10, 25)],
    ]

    for row in data:
        ws.append(row)

    # 열 너비
    for col, width in [('A', 12), ('B', 15), ('C', 15), ('D', 20), ('E', 12), ('F', 15), ('G', 12)]:
        ws.column_dimensions[col].width = width

    output_path = OUTPUT_DIR / "therapists.xlsx"
    wb.save(output_path)
    print(f"✅ 생성: {output_path}")


# ============================================================
# 📌 5. Large Dataset (대용량 테스트용)
# ============================================================

def create_large_dataset_excel():
    """대용량 데이터셋 (1000행)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "LargeDataset"

    # 헤더
    headers = ["ID", "Name", "Phone", "Email", "Amount", "Status", "Created Date"]
    ws.append(headers)

    # 헤더 스타일
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        apply_style(cell, get_header_style())

    # 1000행의 데이터 생성
    for i in range(1, 1001):
        name = f"Person_{i:04d}"
        phone = f"010-{1000 + (i % 9000)}-{5000 + (i % 5000)}"
        email = f"person{i}@example.com"
        amount = 100000 + (i * 1000) % 500000
        status = ["Active", "Pending", "Completed"][i % 3]
        created_date = date(2025, 1, 1) + timedelta(days=i % 180)

        ws.append([i, name, phone, email, amount, status, created_date])

    # 열 너비
    for col, width in [('A', 8), ('B', 15), ('C', 15), ('D', 20), ('E', 12), ('F', 12), ('G', 12)]:
        ws.column_dimensions[col].width = width

    output_path = OUTPUT_DIR / "large_dataset.xlsx"
    wb.save(output_path)
    print(f"✅ 생성: {output_path} (1000행)")


# ============================================================
# 📌 메인 실행
# ============================================================

if __name__ == "__main__":
    print("=" * 60)
    print("📌 Sample Excel Files Generator")
    print("=" * 60)

    create_bookings_excel()
    create_employees_excel()
    create_expenses_excel()
    create_therapists_excel()
    create_large_dataset_excel()

    print("\n" + "=" * 60)
    print("✅ 모든 샘플 파일 생성 완료!")
    print(f"📁 저장 위치: {OUTPUT_DIR}")
    print("=" * 60)

    # 생성된 파일 목록
    print("\n📋 생성된 파일:")
    for file in sorted(OUTPUT_DIR.glob("*.xlsx")):
        file_size = file.stat().st_size
        print(f"  - {file.name} ({file_size:,} bytes)")
