"""
SSS Electronic Contribution Collection List → Excel 변환
Source: KakaoTalk_20260520_055522727_02.jpg (summary)
        KakaoTalk_20260520_055522727_03.jpg (employee list)
Month: September 2024
Company: THE EL INC.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from datetime import date

# ──────────────────────────────────────────
# 이미지에서 읽은 직원 데이터
# No | Name | SS Number | SS | EC | Total
# ──────────────────────────────────────────
EMPLOYEES = [
    ( 1, "ATILLO, MARSING M",        "06-3411494-8",    0.00,    0.00,    0.00),
    ( 2, "BERNANDY, JANICE B",       "08-3411494-8",    0.00,    0.00,    0.00),
    ( 3, "BRIONES, BRENDA V",        "08-2225130-0", 1120.00,   10.00, 1130.00),
    ( 4, "CALUNO, LAURA N",          "06-1897293-0", 1610.00,   10.00, 1620.00),
    ( 5, "CRISTOBAL, EVELYN L",      "08-7879044-6", 1610.00,   10.00, 1620.00),
    ( 6, "DAUDA, JANETTE P",         "08-1817699-4", 1000.00,   10.00, 1010.00),
    ( 7, "DELITO, CRISTY C",         "08-1815444-5",    0.00,    0.00,    0.00),
    ( 8, "DURANO, RUE P",            "09-1394539-3",    0.00,    0.00,    0.00),
    ( 9, "ELSCDON, JANCE D",         "09-3100843-1", 1120.00,   10.00, 1130.00),
    (10, "ESCOBIDO, ADELAIDA J",     "09-1489824-1",    0.00,    0.00,    0.00),
    (11, "ESPINOSA, PHILIP R",       "06-2883734-1", 1610.00,   10.00, 1620.00),
    (12, "ESPINOSA, JOSEPHINE D",    "09-2283939-7", 1610.00,   10.00, 1620.00),
    (13, "ESTOLA, JESSICA J",        "09-2836918-9",    0.00,    0.00,    0.00),
    (14, "FRANCIS, JESSICA J",       "09-2638316-9",    0.00,    0.00,    0.00),
    (15, "GINA, KIMBERLY B",         "08-7233390-0",    0.00,    0.00,    0.00),
    (16, "IQOT, DANA PRINCESS N",    "05-4820011-0", 1610.00,   10.00, 1620.00),
    (17, "IQOT, JANELA",             "05-4930456-6", 1610.00,   10.00, 1620.00),
    (18, "LLABRES, EUGENE D",        "06-2831815-8", 1610.00,   10.00, 1620.00),
    (19, "MANGUBAT, JENNIFER A",     "06-2291634-9", 1610.00,   10.00, 1620.00),
    (20, "MOREVA, WILMA O",          "06-1614545-4",    0.00,    0.00,    0.00),
    (21, "PACALDO, MA. TERESITA B",  "06-2217242-9",    0.00,    0.00,    0.00),
    (22, "PECIA, ARCELYE E",         "34-5589282-8", 1120.00,   10.00, 1130.00),
    (23, "ROSAS, WILFREDA M",        "06-2891027-8", 1120.00,   10.00, 1130.00),
    (24, "SHRIFT, KIMBERLEY ACE L",  "06-2413130-9", 2100.00,   30.00, 2130.00),
    (25, "TAMPAN, FATIMA A",         "06-2777165-5", 1610.00,   10.00, 1620.00),
    (26, "TANSE, ROSELYN J",         "06-2583395-0", 1610.00,   10.00, 1620.00),
    (27, "TEJANO, BRUNA T",          "06-3705050-3", 1810.00,   10.00, 1820.00),
    (28, "TEJANO, JOVELYN T",        "06-2483803-0", 1810.00,   10.00, 1820.00),
    (29, "*(TEJANO, JOVELYN T)",     "06-2483803-0",    0.00,    0.00,    0.00),
]

# ──────────────────────────────────────────
# 스타일 정의
# ──────────────────────────────────────────
BLUE_DARK   = "1F3864"
BLUE_MID    = "2E75B6"
BLUE_LIGHT  = "BDD7EE"
YELLOW_HDR  = "FFF2CC"
ORANGE_HDR  = "F4B942"
GRAY_ROW    = "F5F5F5"
GREEN_TOTAL = "E2EFDA"
RED_ZERO    = "FCE4D6"

thin = Side(style="thin", color="AAAAAA")
med  = Side(style="medium", color="555555")
BORDER_THIN  = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
BORDER_MED   = Border(left=med,   right=med,   top=med,   bottom=med)
BORDER_BOT_M = Border(left=thin,  right=thin,  top=thin,  bottom=med)

def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def cell_style(ws, row, col, value=None,
               bold=False, italic=False, size=10, color="000000",
               fill=None, align_h="left", align_v="center",
               border=BORDER_THIN, fmt=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, italic=italic, size=size, color=color, name="Calibri")
    c.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=wrap)
    if fill:
        c.fill  = hdr_fill(fill)
    if border:
        c.border = border
    if fmt:
        c.number_format = fmt
    return c


# ──────────────────────────────────────────
# 워크북 생성
# ──────────────────────────────────────────
wb = openpyxl.Workbook()

# ══════════════════════════════════════════
# SHEET 1 : Employee Detail
# ══════════════════════════════════════════
ws = wb.active
ws.title = "SSS September 2024"

# --- column widths ---
col_widths = [5, 32, 18, 14, 14, 14]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# --- Row 1-2: Company header ---
ws.merge_cells("A1:F1")
cell_style(ws, 1, 1,
           value="SOCIAL SECURITY SYSTEM — ELECTRONIC CONTRIBUTION COLLECTION LIST",
           bold=True, size=13, color="FFFFFF", fill=BLUE_DARK,
           align_h="center", border=BORDER_MED)
ws.row_dimensions[1].height = 22

ws.merge_cells("A2:F2")
cell_style(ws, 2, 1,
           value="THE EL INC.  |  Airport Road, Pusok, Lapu-Lapu City (Open) Cebu 6015",
           bold=False, size=10, color="333333", fill=BLUE_LIGHT,
           align_h="center", border=BORDER_THIN)

# --- Row 3: Meta info ---
meta = [
    ("Employer SS No:", "80-0148613-3-000"),
    ("Applicable Month:", "September 2024"),
    ("Invoice No:", "5A26006208489"),
]
for col_start, (label, val) in zip([1, 3, 5], meta):
    cell_style(ws, 3, col_start, value=label,   bold=True,  size=9, fill=YELLOW_HDR, align_h="right")
    cell_style(ws, 3, col_start+1, value=val,   bold=False, size=9, fill=YELLOW_HDR, align_h="left")

# --- Row 4: blank spacer ---
ws.row_dimensions[4].height = 6

# --- Row 5: Column headers ---
headers = ["No.", "Name of Employee", "SS Number", "SS\n(PHP)", "EC\n(PHP)", "Total\n(PHP)"]
hdr_aligns = ["center","left","center","center","center","center"]
for col, (h, ha) in enumerate(zip(headers, hdr_aligns), 1):
    cell_style(ws, 5, col, value=h,
               bold=True, size=10, color="FFFFFF",
               fill=BLUE_MID, align_h=ha, align_v="center",
               wrap=True, border=BORDER_MED)
ws.row_dimensions[5].height = 28

# --- Rows 6+: Employee data ---
PESO = '#,##0.00'
for idx, (no, name, ssn, ss, ec, total) in enumerate(EMPLOYEES):
    row = idx + 6
    row_fill = GRAY_ROW if idx % 2 == 0 else "FFFFFF"

    cell_style(ws, row, 1, no,    size=9, align_h="center", fill=row_fill)
    cell_style(ws, row, 2, name,  size=9, align_h="left",   fill=row_fill)
    cell_style(ws, row, 3, ssn,   size=9, align_h="center", fill=row_fill)

    # Highlight zero contributions in light red
    ss_fill  = RED_ZERO if ss == 0  else row_fill
    ec_fill  = RED_ZERO if ec == 0  else row_fill
    tot_fill = RED_ZERO if total == 0 else row_fill

    cell_style(ws, row, 4, ss,    size=9, align_h="right",  fill=ss_fill,  fmt=PESO)
    cell_style(ws, row, 5, ec,    size=9, align_h="right",  fill=ec_fill,  fmt=PESO)
    cell_style(ws, row, 6, total, size=9, align_h="right",  fill=tot_fill, fmt=PESO)
    ws.row_dimensions[row].height = 16

# --- Totals row ---
tot_row = 6 + len(EMPLOYEES)
ss_sum    = sum(e[3] for e in EMPLOYEES)
ec_sum    = sum(e[4] for e in EMPLOYEES)
total_sum = sum(e[5] for e in EMPLOYEES)

ws.merge_cells(f"A{tot_row}:C{tot_row}")
cell_style(ws, tot_row, 1, "SUB-TOTAL",
           bold=True, size=10, color="FFFFFF",
           fill=BLUE_MID, align_h="center", border=BORDER_BOT_M)
cell_style(ws, tot_row, 4, ss_sum,    bold=True, size=10, fill=GREEN_TOTAL, align_h="right", fmt=PESO, border=BORDER_BOT_M)
cell_style(ws, tot_row, 5, ec_sum,    bold=True, size=10, fill=GREEN_TOTAL, align_h="right", fmt=PESO, border=BORDER_BOT_M)
cell_style(ws, tot_row, 6, total_sum, bold=True, size=10, fill=GREEN_TOTAL, align_h="right", fmt=PESO, border=BORDER_BOT_M)
ws.row_dimensions[tot_row].height = 20

# --- Note row ---
note_row = tot_row + 2
ws.merge_cells(f"A{note_row}:F{note_row}")
cell_style(ws, note_row, 1,
           value="⚠  Items highlighted in red (PHP 0.00) indicate no contribution for this period. Please verify with payroll records.",
           italic=True, size=8, color="AA0000", fill="FFF0F0",
           align_h="left", border=None)

# --- Freeze panes ---
ws.freeze_panes = "A6"


# ══════════════════════════════════════════
# SHEET 2 : Summary
# ══════════════════════════════════════════
ws2 = wb.create_sheet("Summary")
ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 22

ws2.merge_cells("A1:B1")
cell_style(ws2, 1, 1, "SSS PAYMENT SUMMARY",
           bold=True, size=14, color="FFFFFF", fill=BLUE_DARK,
           align_h="center", border=BORDER_MED)
ws2.row_dimensions[1].height = 24

summary_data = [
    ("Company",            "THE EL INC."),
    ("Employer SS No.",    "80-0148613-3-000"),
    ("Applicable Month",   "September 2024"),
    ("Invoice No.",        "5A26006208489"),
    ("Generation Date",    "January 04, 2026"),
    ("Due Date",           "DUE IMMEDIATELY"),
    ("",                   ""),
    ("Total Employees",    29),
    ("Employees w/ SS >0", sum(1 for e in EMPLOYEES if e[3] > 0)),
    ("",                   ""),
    ("SS Contribution",    ss_sum),
    ("EC Contribution",    ec_sum),
    ("TOTAL AMOUNT DUE",   total_sum),
]

for i, (label, val) in enumerate(summary_data, 2):
    is_total = label == "TOTAL AMOUNT DUE"
    fill_l = BLUE_LIGHT if not is_total else ORANGE_HDR
    fill_v = "FFFFFF"   if not is_total else YELLOW_HDR
    fmt    = PESO if isinstance(val, float) else None

    cell_style(ws2, i, 1, label,
               bold=is_total, size=10, fill=fill_l, align_h="left")
    cell_style(ws2, i, 2, val,
               bold=is_total, size=10, fill=fill_v, align_h="right", fmt=fmt)


# ══════════════════════════════════════════
# Save
# ══════════════════════════════════════════
OUT = r"e:\elspa\sss\SSS_September_2024_THE_EL_INC.xlsx"
wb.save(OUT)
print(f"✅  Excel saved → {OUT}")
print(f"    Employees : {len(EMPLOYEES)}")
print(f"    SS Total  : PHP {ss_sum:,.2f}")
print(f"    EC Total  : PHP {ec_sum:,.2f}")
print(f"    Grand Total: PHP {total_sum:,.2f}")
